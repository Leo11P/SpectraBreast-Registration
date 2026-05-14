"""
SPECTRABREAST — BATCH RUNNER
=============================
Esegue la pipeline di registrazione su PIU' coppie (mesh + aruco_json)
in un solo comando, scandendo data/input/ alla ricerca di cartelle che
matchano i pattern dei sample (default: "SAMPLE1*" e "SAMPLE2*").

Per ogni cartella trovata:
  - cerca al suo interno  surface_mesh.ply  e  aruco_markers_3d.json
  - usa l'HSI .hdr definito dalla mappa sample_hsi_map nel config
  - esegue run_full_pipeline() in output_dir/<nome_cartella_sample>/

Al termine produce un Excel riepilogo unico con i risultati di tutte le coppie.

Mutuamente esclusivo con sweep: se batch.enabled e' true, sweep viene ignorato.
"""

from __future__ import annotations

import fnmatch
import os
import time
import traceback
from typing import Optional

import numpy as np

from spectrabreast.pipeline import run_full_pipeline, load_mesh


# =============================================================================
# Stats helpers (riusiamo la logica di sweep.py)
# =============================================================================

def _stats(arr) -> tuple[float, float, int]:
    if arr is None:
        return float('nan'), float('nan'), 0
    a = np.asarray(arr, dtype=np.float64)
    if a.size == 0:
        return float('nan'), float('nan'), 0
    v = a[~np.isnan(a)]
    if v.size == 0:
        return float('nan'), float('nan'), 0
    return float(v.mean()), float(np.median(v)), int(v.size)


def _px_per_mm_from_data(data_hsi: dict, marker_side_mm: float) -> float:
    sides = []
    for corners in data_hsi.values():
        c = np.asarray(corners, dtype=np.float32)
        sides.append(float(np.mean([
            np.linalg.norm(c[(j + 1) % 4] - c[j]) for j in range(4)
        ])))
    return float(np.mean(sides)) / marker_side_mm if sides else 1.0


# =============================================================================
# Discovery delle coppie sample
# =============================================================================

def _resolve_sample_hsi_path(hsi_value: str, base_dir: str) -> str:
    """
    Risolve un path HSI dalla config:
      - se assoluto -> usato cosi' com'e'
      - se relativo -> risolto rispetto a base_dir (root del progetto)
    """
    if os.path.isabs(hsi_value):
        return hsi_value
    return os.path.join(base_dir, hsi_value)


def discover_sample_pairs(
    input_dir: str,
    sample_hsi_map: dict,
    base_dir: str,
    mesh_filename: str = 'surface_mesh.ply',
    aruco_filename: str = 'aruco_markers_3d.json',
) -> list[dict]:
    """
    Scansiona input_dir alla ricerca di cartelle che matchano i pattern
    delle chiavi di sample_hsi_map (es. "SAMPLE1*", "SAMPLE2*").

    Per ogni cartella trovata verifica la presenza di mesh + aruco_json e
    determina l'HSI associato dalla mappa.

    Returns
    -------
    pairs : list of dict, ciascuno con
        'sample_dir_name' : nome della cartella (es. "SAMPLE1_run1")
        'sample_pattern'  : pattern matchato (es. "SAMPLE1*")
        'mesh_path'       : path assoluto al .ply
        'aruco_path'      : path assoluto al .json
        'hsi_hdr_path'    : path assoluto al .hdr (dalla mappa)
        'output_subdir'   : nome della sottocartella di output da usare
        'sample_name'     : nome usato come prefisso file (= sample_dir_name)
    """
    if not os.path.isdir(input_dir):
        raise FileNotFoundError(
            f"[batch] Cartella input non trovata: {input_dir}"
        )

    if not sample_hsi_map:
        raise ValueError(
            "[batch] sample_hsi_map vuoto. Definisci almeno un pattern "
            "(es. SAMPLE1*) -> percorso HSI."
        )

    all_entries = sorted(os.listdir(input_dir))
    pairs: list[dict] = []
    skipped: list[tuple[str, str]] = []

    for entry in all_entries:
        full_path = os.path.join(input_dir, entry)
        if not os.path.isdir(full_path):
            continue

        # Cerca un pattern che matcha
        matched_pattern = None
        for pattern in sample_hsi_map.keys():
            if fnmatch.fnmatch(entry, pattern):
                matched_pattern = pattern
                break
        if matched_pattern is None:
            continue   # non rientra nei pattern dei sample da processare

        mesh_path  = os.path.join(full_path, mesh_filename)
        aruco_path = os.path.join(full_path, aruco_filename)

        missing = []
        if not os.path.isfile(mesh_path):
            missing.append(mesh_filename)
        if not os.path.isfile(aruco_path):
            missing.append(aruco_filename)
        if missing:
            skipped.append((entry, f"manca: {', '.join(missing)}"))
            continue

        hsi_hdr_raw  = sample_hsi_map[matched_pattern]
        hsi_hdr_path = _resolve_sample_hsi_path(hsi_hdr_raw, base_dir)
        if not os.path.isfile(hsi_hdr_path):
            skipped.append((entry, f"HSI mancante: {hsi_hdr_path}"))
            continue

        pairs.append({
            'sample_dir_name': entry,
            'sample_pattern' : matched_pattern,
            'mesh_path'      : mesh_path,
            'aruco_path'     : aruco_path,
            'hsi_hdr_path'   : hsi_hdr_path,
            'output_subdir'  : entry,
            'sample_name'    : entry,
        })

    # Report
    print(f"\n[batch] Scansionata cartella: {input_dir}")
    print(f"[batch] Pattern attivi      : {list(sample_hsi_map.keys())}")
    print(f"[batch] Coppie valide       : {len(pairs)}")
    for p in pairs:
        print(f"  + {p['sample_dir_name']:30s}  "
              f"(pattern={p['sample_pattern']})")
        print(f"      mesh : {os.path.relpath(p['mesh_path'], base_dir)}")
        print(f"      aruco: {os.path.relpath(p['aruco_path'], base_dir)}")
        print(f"      hsi  : {os.path.relpath(p['hsi_hdr_path'], base_dir)}")
    if skipped:
        print(f"\n[batch] Cartelle ignorate: {len(skipped)}")
        for name, reason in skipped:
            print(f"  - {name:30s}  {reason}")

    return pairs


# =============================================================================
# Singola run nel batch
# =============================================================================

def _run_single_sample(
    pair: dict,
    cfg: dict,
    aruco_dict_cv: int,
    torch_device: Optional[str],
    use_torch_render: bool,
    render_fn,
) -> dict:
    """
    Esegue la pipeline completa per UNA coppia mesh+aruco.

    Calcola i render GPU una sola volta (pc + reg) prima della pipeline,
    esattamente come fa main.py per la run singola.
    """
    sample_name = pair['sample_name']
    output_dir  = os.path.join(cfg['paths']['output_dir'], pair['output_subdir'])
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "=" * 68)
    print(f"  BATCH RUN  —  {sample_name}")
    print(f"  Pattern : {pair['sample_pattern']}")
    print(f"  Output  : {output_dir}")
    print("=" * 68)

    save_pointcloud = cfg['export'].get('save_pointcloud', True)
    save_images     = cfg['export'].get('save_images', True)

    res_pc  = cfg['render']['resolution_mm_per_px']
    res_reg = cfg['render'].get('resolution_reg_mm_per_px', res_pc)
    dual    = abs(res_reg - res_pc) > 1e-6

    # ── Pre-calcolo render GPU (come in main.py) ─────────────────────────────
    precomputed_render     = None
    precomputed_render_reg = None

    if use_torch_render and render_fn is not None:
        mesh = load_mesh(pair['mesh_path'], scale_m_to_mm=True)

        print(f"\n[batch] Render PC ({res_pc} mm/px) su {torch_device}...")
        t0 = time.time()
        precomputed_render = render_fn(
            mesh,
            resolution_mm_per_px=res_pc,
            margin_mm=cfg['render']['margin_mm'],
            device=torch_device,
        )
        print(f"[batch] Render PC completato in {time.time()-t0:.1f}s")

        if dual:
            print(f"\n[batch] Render REG ({res_reg} mm/px) su {torch_device}...")
            t0 = time.time()
            precomputed_render_reg = render_fn(
                mesh,
                resolution_mm_per_px=res_reg,
                margin_mm=cfg['render']['margin_mm'],
                device=torch_device,
            )
            print(f"[batch] Render REG completato in {time.time()-t0:.1f}s")
        else:
            precomputed_render_reg = precomputed_render
            print("[batch] reg == pc — render condiviso.")

    # ── Pipeline ─────────────────────────────────────────────────────────────
    t_pipe = time.time()
    result = run_full_pipeline(
        hsi_hdr_path             = pair['hsi_hdr_path'],
        mesh_path                = pair['mesh_path'],
        aruco_json_path          = pair['aruco_path'],
        output_dir               = output_dir,
        hsi_extraction_method    = cfg['registration']['hsi_extraction_method'],
        aruco_dict_type          = aruco_dict_cv,
        suspicious_pixels_hsi    = None,
        render_resolution_mm     = res_pc,
        render_resolution_reg_mm = res_reg,
        render_margin_mm         = cfg['render']['margin_mm'],
        marker_side_mm           = cfg['registration']['marker_side_mm'],
        use_subpix               = cfg['registration'].get('use_subpix', True),
        subpix_winsize           = cfg['registration'].get('subpix_winsize', 5),
        border_px                = cfg['pointcloud']['border_px'],
        reflectance_norm         = cfg['pointcloud']['reflectance_norm'],
        pc_chunk_size            = cfg['pointcloud'].get('pc_chunk_size', 100_000),
        export_ply_file          = cfg['export']['ply'] if save_pointcloud else False,
        export_npy_file          = cfg['export']['npy'] if save_pointcloud else False,
        export_csv_file          = cfg['export']['csv'] if save_pointcloud else False,
        csv_max_points           = cfg['export']['csv_max_points'],
        save_pointcloud          = save_pointcloud,
        save_images              = save_images,
        sample_name              = sample_name,
        precomputed_render       = precomputed_render,
        precomputed_render_reg   = precomputed_render_reg,
    )
    elapsed = time.time() - t_pipe

    # ── Estrazione metriche per il riepilogo ─────────────────────────────────
    err2d_px      = result.get('err2d_px')
    err3d_dict    = result.get('err3d_dict')
    err3d_dict_pc = result.get('err3d_dict_pc')
    data_hsi      = result.get('data_hsi')

    if data_hsi:
        px_per_mm = _px_per_mm_from_data(data_hsi, cfg['registration']['marker_side_mm'])
    else:
        px_per_mm = float('nan')

    if err2d_px is not None:
        err2d = np.asarray(err2d_px, dtype=np.float64)
        m2px, med2px, n2 = _stats(err2d)
        if not np.isnan(px_per_mm) and px_per_mm > 0:
            err2d_mm = err2d / px_per_mm
            m2mm, med2mm, _ = _stats(err2d_mm)
        else:
            m2mm = med2mm = float('nan')
    else:
        m2px = med2px = m2mm = med2mm = float('nan')
        n2 = 0

    if err3d_dict is not None:
        m_rb_bil, md_rb_bil, n_rb_bil = _stats(err3d_dict.get('bilinear'))
        m_rb_bic, md_rb_bic, n_rb_bic = _stats(err3d_dict.get('bicubic'))
    else:
        m_rb_bil = md_rb_bil = m_rb_bic = md_rb_bic = float('nan')
        n_rb_bil = n_rb_bic = 0

    if err3d_dict_pc is not None:
        m_pc_bil, md_pc_bil, n_pc_bil = _stats(err3d_dict_pc.get('bilinear'))
        m_pc_bic, md_pc_bic, n_pc_bic = _stats(err3d_dict_pc.get('bicubic'))
    else:
        m_pc_bil, md_pc_bil, n_pc_bil = m_rb_bil, md_rb_bil, n_rb_bil
        m_pc_bic, md_pc_bic, n_pc_bic = m_rb_bic, md_rb_bic, n_rb_bic

    n_pts = result.get('n_valid', 0)
    bands = result.get('bands', 0)

    return {
        'sample'                        : sample_name,
        'pattern'                       : pair['sample_pattern'],
        'output_subdir'                 : pair['output_subdir'],
        'res_reg_mm_pix'                : res_reg,
        'res_pc_mm_pix'                 : res_pc,
        '2D_mean_px'                    : m2px,
        '2D_median_px'                  : med2px,
        '2D_mean_mm'                    : m2mm,
        '2D_median_mm'                  : med2mm,
        '3D_REG_bilinear_mean_mm'       : m_rb_bil,
        '3D_REG_bilinear_median_mm'     : md_rb_bil,
        '3D_REG_bicubic_mean_mm'        : m_rb_bic,
        '3D_REG_bicubic_median_mm'      : md_rb_bic,
        '3D_PC_bilinear_mean_mm'        : m_pc_bil,
        '3D_PC_bilinear_median_mm'      : md_pc_bil,
        '3D_PC_bicubic_mean_mm'         : m_pc_bic,
        '3D_PC_bicubic_median_mm'       : md_pc_bic,
        'n_corners_2D'                  : n2,
        'n_corners_3D_REG_bilinear'     : n_rb_bil,
        'n_corners_3D_REG_bicubic'      : n_rb_bic,
        'n_corners_3D_PC_bilinear'      : n_pc_bil,
        'n_corners_3D_PC_bicubic'       : n_pc_bic,
        'n_points_cloud'                : n_pts,
        'n_bands'                       : bands,
        'elapsed_s'                     : round(elapsed, 2),
        'status'                        : 'ok',
    }


def _empty_row(pair: dict, res_reg: float, res_pc: float, status: str) -> dict:
    nan = float('nan')
    return {
        'sample'                        : pair['sample_name'],
        'pattern'                       : pair['sample_pattern'],
        'output_subdir'                 : pair['output_subdir'],
        'res_reg_mm_pix'                : res_reg,
        'res_pc_mm_pix'                 : res_pc,
        '2D_mean_px'                    : nan, '2D_median_px'                  : nan,
        '2D_mean_mm'                    : nan, '2D_median_mm'                  : nan,
        '3D_REG_bilinear_mean_mm'       : nan, '3D_REG_bilinear_median_mm'     : nan,
        '3D_REG_bicubic_mean_mm'        : nan, '3D_REG_bicubic_median_mm'      : nan,
        '3D_PC_bilinear_mean_mm'        : nan, '3D_PC_bilinear_median_mm'      : nan,
        '3D_PC_bicubic_mean_mm'         : nan, '3D_PC_bicubic_median_mm'       : nan,
        'n_corners_2D'                  : 0,
        'n_corners_3D_REG_bilinear'     : 0, 'n_corners_3D_REG_bicubic'      : 0,
        'n_corners_3D_PC_bilinear'      : 0, 'n_corners_3D_PC_bicubic'       : 0,
        'n_points_cloud'                : 0,
        'n_bands'                       : 0,
        'elapsed_s'                     : 0.0,
        'status'                        : status,
    }


# =============================================================================
# Excel riepilogo batch
# =============================================================================

def _write_batch_summary_xlsx(rows: list[dict], output_path: str) -> None:
    """
    Scrive un Excel con una riga per coppia sample processata.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Batch Summary'

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', start_color='305496')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(border_style='thin', color='BFBFBF')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    norm_font = Font(name='Calibri', size=10)

    sample_fill = PatternFill('solid', start_color='FFF2CC')
    res_fill    = PatternFill('solid', start_color='F8CBAD')
    err2d_fill  = PatternFill('solid', start_color='DDEBF7')
    reg_fill    = PatternFill('solid', start_color='D9E1F2')
    pc_fill     = PatternFill('solid', start_color='E2EFDA')
    meta_fill   = PatternFill('solid', start_color='F2F2F2')
    err_fill    = PatternFill('solid', start_color='F4B084')

    ws.cell(row=1, column=1, value=f"Batch Summary — {len(rows)} sample").font = \
        Font(bold=True, size=14, color='1F4E78')
    ws.cell(row=2, column=1,
            value=f"Generato: {time.strftime('%Y-%m-%d %H:%M:%S')}").font = \
        Font(italic=True, color='595959')

    columns: list[tuple[str, str, PatternFill]] = [
        ('sample',                    'Sample',                        sample_fill),
        ('pattern',                   'Pattern',                       sample_fill),
        ('output_subdir',             'Output dir',                    sample_fill),
        ('res_reg_mm_pix',            'res_reg\n(mm/px)',              res_fill),
        ('res_pc_mm_pix',             'res_pc\n(mm/px)',               res_fill),
        ('2D_mean_px',                '2D mean\n(px)',                 err2d_fill),
        ('2D_median_px',              '2D median\n(px)',               err2d_fill),
        ('2D_mean_mm',                '2D mean\n(mm)',                 err2d_fill),
        ('2D_median_mm',              '2D median\n(mm)',               err2d_fill),
        ('3D_REG_bilinear_mean_mm',   '3D REG bilinear\nmean (mm)',    reg_fill),
        ('3D_REG_bilinear_median_mm', '3D REG bilinear\nmedian (mm)',  reg_fill),
        ('3D_REG_bicubic_mean_mm',    '3D REG bicubic\nmean (mm)',     reg_fill),
        ('3D_REG_bicubic_median_mm',  '3D REG bicubic\nmedian (mm)',   reg_fill),
        ('3D_PC_bilinear_mean_mm',    '3D PC bilinear\nmean (mm)',     pc_fill),
        ('3D_PC_bilinear_median_mm',  '3D PC bilinear\nmedian (mm)',   pc_fill),
        ('3D_PC_bicubic_mean_mm',     '3D PC bicubic\nmean (mm)',      pc_fill),
        ('3D_PC_bicubic_median_mm',   '3D PC bicubic\nmedian (mm)',    pc_fill),
        ('n_corners_2D',              'N corners\n2D',                 meta_fill),
        ('n_corners_3D_REG_bilinear', 'N corners\n3D REG bil',         meta_fill),
        ('n_corners_3D_REG_bicubic',  'N corners\n3D REG bic',         meta_fill),
        ('n_corners_3D_PC_bilinear',  'N corners\n3D PC bil',          meta_fill),
        ('n_corners_3D_PC_bicubic',   'N corners\n3D PC bic',          meta_fill),
        ('n_points_cloud',            'N points\ncloud',               meta_fill),
        ('n_bands',                   'N bands',                       meta_fill),
        ('elapsed_s',                 'Elapsed\n(s)',                  meta_fill),
        ('status',                    'Status',                        meta_fill),
    ]

    HEADER_ROW = 4
    for ci, (_, label, _) in enumerate(columns, start=1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = border
    ws.row_dimensions[HEADER_ROW].height = 36

    def _fmt(v):
        return 'N/A' if isinstance(v, float) and np.isnan(v) else v

    for row_i, row in enumerate(rows, start=HEADER_ROW + 1):
        is_err = row.get('status') != 'ok'
        for ci, (key, _, fill) in enumerate(columns, start=1):
            val = row.get(key, '')
            c   = ws.cell(row=row_i, column=ci, value=_fmt(val))
            c.alignment = center
            c.border    = border
            c.font      = norm_font
            c.fill      = err_fill if is_err else fill
            if isinstance(val, float) and not np.isnan(val):
                if key.endswith(('_mm', '_px')) or key == 'elapsed_s':
                    c.number_format = '0.0000'

    for ci, (_, label, _) in enumerate(columns, start=1):
        max_len = max(len(line) for line in label.split('\n'))
        ws.column_dimensions[get_column_letter(ci)].width = max(max_len + 3, 12)

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=4)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n[batch] Excel riepilogo salvato -> {output_path}")


# =============================================================================
# Entry point
# =============================================================================

def run_batch(
    cfg: dict,
    aruco_dict_cv: int,
    torch_device: Optional[str],
    use_torch_render: bool,
    base_dir: str,
    render_fn=None,
) -> str:
    """
    Esegue il batch su tutte le coppie discoverte e ritorna il path dell'Excel.

    Parameters
    ----------
    cfg              : config caricato (con path gia' risolti)
    aruco_dict_cv    : costante OpenCV del dizionario ArUco
    torch_device     : device torch (es. 'cuda', 'cpu')
    use_torch_render : True se il render GPU e' disponibile
    base_dir         : root del progetto (per risolvere path relativi degli HSI)
    render_fn        : funzione render_orthographic_topview_gpu (o None)
    """
    batch_cfg = cfg.get('batch', {}) or {}

    input_dir = batch_cfg.get('input_dir')
    if input_dir is None:
        # fallback: usa la cartella che contiene la mesh dichiarata nel config
        input_dir = os.path.dirname(os.path.dirname(cfg['paths']['mesh']))
    if not os.path.isabs(input_dir):
        input_dir = os.path.join(base_dir, input_dir)

    sample_hsi_map = batch_cfg.get('sample_hsi_map', {}) or {}
    mesh_filename  = batch_cfg.get('mesh_filename',  'surface_mesh.ply')
    aruco_filename = batch_cfg.get('aruco_filename', 'aruco_markers_3d.json')

    print("\n" + "=" * 68)
    print("  BATCH MODE")
    print("=" * 68)
    print(f"  Input dir       : {input_dir}")
    print(f"  Mesh filename   : {mesh_filename}")
    print(f"  Aruco filename  : {aruco_filename}")
    print(f"  Sample patterns : {list(sample_hsi_map.keys())}")

    pairs = discover_sample_pairs(
        input_dir      = input_dir,
        sample_hsi_map = sample_hsi_map,
        base_dir       = base_dir,
        mesh_filename  = mesh_filename,
        aruco_filename = aruco_filename,
    )

    if not pairs:
        print("\n[batch] Nessuna coppia valida trovata. Esco.")
        return ''

    base_out = cfg['paths']['output_dir']
    os.makedirs(base_out, exist_ok=True)

    res_pc  = cfg['render']['resolution_mm_per_px']
    res_reg = cfg['render'].get('resolution_reg_mm_per_px', res_pc)

    rows: list[dict] = []
    t_total = time.time()

    for i, pair in enumerate(pairs):
        print(f"\n[batch] Progresso: {i + 1}/{len(pairs)}  "
              f"-> {pair['sample_dir_name']}")
        try:
            row = _run_single_sample(
                pair             = pair,
                cfg              = cfg,
                aruco_dict_cv    = aruco_dict_cv,
                torch_device     = torch_device,
                use_torch_render = use_torch_render,
                render_fn        = render_fn,
            )
        except Exception as e:
            print(f"\n[batch] ERRORE su {pair['sample_dir_name']}: {e}")
            traceback.print_exc()
            row = _empty_row(pair, res_reg, res_pc,
                             status=f'error: {type(e).__name__}: {e}')
        rows.append(row)

    total_elapsed = time.time() - t_total

    # ── Excel riepilogo ──────────────────────────────────────────────────────
    out_path = os.path.join(base_out, 'batch_summary.xlsx')
    _write_batch_summary_xlsx(rows, out_path)

    # ── Riepilogo finale a terminale ─────────────────────────────────────────
    print("\n" + "=" * 68)
    print("BATCH COMPLETATO")
    print("=" * 68)
    n_ok = sum(1 for r in rows if r['status'] == 'ok')
    print(f"  Coppie processate : {len(rows)}")
    print(f"  OK / Falliti      : {n_ok} / {len(rows) - n_ok}")
    print(f"  Tempo totale      : {total_elapsed:.1f}s  ({total_elapsed/60:.1f} min)")
    print(f"  Output dir        : {base_out}")
    print(f"  Excel riepilogo   : {out_path}")
    print("=" * 68)

    return out_path
