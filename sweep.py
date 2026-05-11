"""
SPECTRABREAST — RESOLUTION SWEEP
=================================
Esegue la pipeline di registrazione su una lista di coppie
(resolution_reg_mm_per_px, resolution_pc_mm_per_px) lette da config.yaml
e produce UN SOLO file Excel riepilogativo nella output_dir.

Ottimizzazione render:
  Prima di iniziare il loop sulle coppie, vengono raccolti tutti i valori
  UNICI di risoluzione (sia da res_reg che da res_pc) e per ognuno viene
  calcolato UN solo render GPU. Le run dello sweep poi riusano i render
  dalla cache, evitando di ricalcolare lo stesso render piu' volte quando
  piu' coppie condividono lo stesso valore.

Salvataggio best result:
  Se nel config export.save_pointcloud: true o export.save_images: true,
  al termine dello sweep viene rieseguita la coppia con il 3D error assoluto
  piu' basso (min tra tutte le colonne 3D mean: REG bilinear, REG bicubic,
  PC bilinear, PC bicubic) con i flag di salvataggio attivi.
  L'output viene scritto direttamente in output_dir con il prefisso del sample.
  Nell'Excel la riga migliore e' evidenziata in verde; la cella specifica che
  ha determinato il best (il valore minimo assoluto) e' evidenziata in arancio.

Durante il loop principale:
  - save_pointcloud  -> forzato False
  - save_images      -> forzato False
  - Excel per-run    -> scritto in _sweep_tmp/run_N/ (rimosso al termine)
  - SOLO l'Excel riepilogativo finale viene scritto in output_dir.
"""

from __future__ import annotations

import os
import shutil
import time
import traceback
from typing import Optional

import numpy as np

from spectrabreast.pipeline import run_full_pipeline, load_mesh

TORCH_AVAILABLE = True  # placeholder; la disponibilita' reale e' gestita da main.py

# Colonne 3D mean usate per determinare la coppia migliore
_3D_ERROR_KEYS = [
    '3D_REG_bilinear_mean_mm',
    '3D_REG_bicubic_mean_mm',
    '3D_PC_bilinear_mean_mm',
    '3D_PC_bicubic_mean_mm',
]


# =============================================================================
# Helper: statistiche robuste su array con NaN
# =============================================================================

def _stats(arr) -> tuple[float, float, int]:
    if arr is None:
        return float('nan'), float('nan'), 0
    a = np.asarray(arr, dtype=np.float64)
    if a.size == 0:
        return float('nan'), float('nan'), 0
    v = a[~np.isnan(a)]
    if v.size == 0:
        return float('nan'), float('nan'), 0
    return float(v.mean()), float(np.median(v)), int(v.size)


def _px_per_mm_from_data(data_hsi: dict, marker_side_mm: float) -> float:
    sides = []
    for corners in data_hsi.values():
        c = np.asarray(corners, dtype=np.float32)
        sides.append(float(np.mean([
            np.linalg.norm(c[(j + 1) % 4] - c[j]) for j in range(4)
        ])))
    return float(np.mean(sides)) / marker_side_mm if sides else 1.0


def _best_3d_val(row: dict) -> float:
    """Ritorna il minimo valore 3D mean valido della riga."""
    vals = [row.get(k, float('nan')) for k in _3D_ERROR_KEYS]
    valid = [v for v in vals if not np.isnan(v)]
    return min(valid) if valid else float('nan')


# =============================================================================
# Parsing / validazione delle coppie dal config
# =============================================================================

def parse_pairs(sweep_cfg: dict) -> list[tuple[float, float]]:
    raw = sweep_cfg.get('resolution_pairs')
    if raw is None:
        raise ValueError("sweep.resolution_pairs non trovato in config.yaml.")
    if not isinstance(raw, (list, tuple)) or len(raw) == 0:
        raise ValueError("sweep.resolution_pairs deve essere una lista NON vuota.")

    pairs: list[tuple[float, float]] = []
    for i, item in enumerate(raw):
        if not isinstance(item, (list, tuple)) or len(item) != 2:
            raise ValueError(
                f"sweep.resolution_pairs[{i}]: atteso [reg, pc] (2 numeri), "
                f"ricevuto {item!r}"
            )
        try:
            r = float(item[0])
            p = float(item[1])
        except (TypeError, ValueError) as e:
            raise ValueError(
                f"sweep.resolution_pairs[{i}]: valori non numerici -> {item!r}"
            ) from e
        if r <= 0 or p <= 0:
            raise ValueError(
                f"sweep.resolution_pairs[{i}]: risoluzioni devono essere > 0, "
                f"ricevuto reg={r}, pc={p}"
            )
        pairs.append((r, p))
    return pairs


# =============================================================================
# Cache dei render
# =============================================================================

_RES_TOL = 1e-6


def _quantize_res(r: float) -> float:
    return round(float(r), 9)


def _collect_unique_resolutions(pairs: list[tuple[float, float]]) -> list[float]:
    seen: dict[float, float] = {}
    for r, p in pairs:
        for v in (r, p):
            k = _quantize_res(v)
            if k not in seen:
                seen[k] = v
    return sorted(seen.values(), reverse=True)


def _precompute_renders(
    mesh,
    unique_resolutions: list[float],
    margin_mm: float,
    torch_device: Optional[str],
    render_fn,
) -> dict[float, tuple]:
    if render_fn is None:
        raise RuntimeError("render_fn e' None: impossibile pre-calcolare i render.")

    cache: dict[float, tuple] = {}
    n = len(unique_resolutions)
    print(f"\n[Sweep][Cache] Pre-calcolo di {n} render unici...")

    total_t0 = time.time()
    for i, res in enumerate(unique_resolutions):
        key = _quantize_res(res)
        print(f"\n[Sweep][Cache] Render {i+1}/{n}: res = {res} mm/px su {torch_device}")
        t0 = time.time()
        render_tuple = render_fn(
            mesh,
            resolution_mm_per_px=res,
            margin_mm=margin_mm,
            device=torch_device,
        )
        cache[key] = render_tuple
        print(f"[Sweep][Cache] Render {i+1}/{n} completato in {time.time()-t0:.1f}s")

    print(f"\n[Sweep][Cache] Tutti i {n} render pronti in {time.time()-total_t0:.1f}s")
    return cache


def _get_render(cache: dict[float, tuple], res: float) -> tuple:
    key = _quantize_res(res)
    if key not in cache:
        for k, v in cache.items():
            if abs(k - key) < _RES_TOL:
                return v
        raise KeyError(
            f"Render per risoluzione {res} mm/px non trovato in cache. "
            f"Chiavi disponibili: {sorted(cache.keys())}"
        )
    return cache[key]


# =============================================================================
# Singola run dello sweep
# =============================================================================

def _run_single_pair(
    cfg: dict,
    aruco_dict_cv: int,
    res_reg: float,
    res_pc: float,
    torch_device: Optional[str],
    use_torch_render: bool,
    tmp_dir: str,
    render_cache: Optional[dict[float, tuple]],
    save_pointcloud_override: bool = False,
    save_images_override: bool = False,
) -> dict:
    """
    Esegue run_full_pipeline per una coppia.

    save_pointcloud_override / save_images_override permettono al re-run
    della coppia migliore di attivare il salvataggio senza toccare il config.
    Nel loop principale entrambi sono False.
    """
    print(f"\n{'=' * 68}")
    print(f"  SWEEP RUN  —  res_reg = {res_reg} mm/px   res_pc = {res_pc} mm/px")
    if save_pointcloud_override or save_images_override:
        print(f"  [BEST RUN — salvataggio attivo: "
              f"PC={save_pointcloud_override}  IMG={save_images_override}]")
    print(f"{'=' * 68}")
    t0 = time.time()

    os.makedirs(tmp_dir, exist_ok=True)
    dual = abs(res_reg - res_pc) > 1e-6

    # ── Render dalla cache ───────────────────────────────────────────────────
    precomputed_render = None
    precomputed_render_reg = None

    if use_torch_render and render_cache is not None:
        precomputed_render = _get_render(render_cache, res_pc)
        print(f"[Sweep][Cache] Render PC ({res_pc} mm/px) recuperato dalla cache.")
        if dual:
            precomputed_render_reg = _get_render(render_cache, res_reg)
            print(f"[Sweep][Cache] Render REG ({res_reg} mm/px) recuperato dalla cache.")
        else:
            precomputed_render_reg = precomputed_render
            print("[Sweep][Cache] reg == pc — render condiviso.")

    # ── Pipeline ─────────────────────────────────────────────────────────────
    result = run_full_pipeline(
        hsi_hdr_path             = cfg['paths']['hsi_hdr'],
        mesh_path                = cfg['paths']['mesh'],
        aruco_json_path          = cfg['paths']['aruco_json'],
        output_dir               = tmp_dir,
        hsi_extraction_method    = cfg['registration']['hsi_extraction_method'],
        aruco_dict_type          = aruco_dict_cv,
        suspicious_pixels_hsi    = None,
        render_resolution_mm     = res_pc,
        render_resolution_reg_mm = res_reg,
        render_margin_mm         = cfg['render']['margin_mm'],
        marker_side_mm           = cfg['registration']['marker_side_mm'],
        use_subpix               = cfg['registration'].get('use_subpix', True),
        subpix_winsize           = cfg['registration'].get('subpix_winsize', 5),
        border_px                = cfg['pointcloud']['border_px'],
        reflectance_norm         = cfg['pointcloud']['reflectance_norm'],
        pc_chunk_size            = cfg['pointcloud'].get('pc_chunk_size', 100_000),
        export_ply_file          = cfg['export']['ply']  if save_pointcloud_override else False,
        export_npy_file          = cfg['export']['npy']  if save_pointcloud_override else False,
        export_csv_file          = cfg['export']['csv']  if save_pointcloud_override else False,
        csv_max_points           = cfg['export']['csv_max_points'],
        save_pointcloud          = save_pointcloud_override,
        save_images              = save_images_override,
        sample_name              = cfg['sample_name'],
        precomputed_render       = precomputed_render,
        precomputed_render_reg   = precomputed_render_reg,
    )

    # ── Estrazione metriche ──────────────────────────────────────────────────
    err2d_px      = result.get('err2d_px')
    err3d_dict    = result.get('err3d_dict')
    err3d_dict_pc = result.get('err3d_dict_pc')
    data_hsi      = result.get('data_hsi')

    if data_hsi:
        px_per_mm = _px_per_mm_from_data(data_hsi, cfg['registration']['marker_side_mm'])
    else:
        px_per_mm = float('nan')

    if err2d_px is not None:
        err2d = np.asarray(err2d_px, dtype=np.float64)
        m2px, med2px, n2 = _stats(err2d)
        if not np.isnan(px_per_mm) and px_per_mm > 0:
            err2d_mm = err2d / px_per_mm
            m2mm, med2mm, _ = _stats(err2d_mm)
        else:
            m2mm, med2mm = float('nan'), float('nan')
    else:
        m2px = med2px = m2mm = med2mm = float('nan')
        n2 = 0

    if err3d_dict is not None:
        m_rb_bil, md_rb_bil, n_rb_bil = _stats(err3d_dict.get('bilinear'))
        m_rb_bic, md_rb_bic, n_rb_bic = _stats(err3d_dict.get('bicubic'))
    else:
        m_rb_bil = md_rb_bil = m_rb_bic = md_rb_bic = float('nan')
        n_rb_bil = n_rb_bic = 0

    if err3d_dict_pc is not None:
        m_pc_bil, md_pc_bil, n_pc_bil = _stats(err3d_dict_pc.get('bilinear'))
        m_pc_bic, md_pc_bic, n_pc_bic = _stats(err3d_dict_pc.get('bicubic'))
    else:
        m_pc_bil, md_pc_bil, n_pc_bil = m_rb_bil, md_rb_bil, n_rb_bil
        m_pc_bic, md_pc_bic, n_pc_bic = m_rb_bic, md_rb_bic, n_rb_bic

    elapsed = time.time() - t0
    print(f"\n[Sweep] Run completata in {elapsed:.1f}s")

    return {
        'res_reg_mm_pix'                : res_reg,
        'res_pc_mm_pix'                 : res_pc,
        '2D_mean_px'                    : m2px,
        '2D_median_px'                  : med2px,
        '2D_mean_mm'                    : m2mm,
        '2D_median_mm'                  : med2mm,
        '3D_REG_bilinear_mean_mm'       : m_rb_bil,
        '3D_REG_bilinear_median_mm'     : md_rb_bil,
        '3D_REG_bicubic_mean_mm'        : m_rb_bic,
        '3D_REG_bicubic_median_mm'      : md_rb_bic,
        '3D_PC_bilinear_mean_mm'        : m_pc_bil,
        '3D_PC_bilinear_median_mm'      : md_pc_bil,
        '3D_PC_bicubic_mean_mm'         : m_pc_bic,
        '3D_PC_bicubic_median_mm'       : md_pc_bic,
        'n_corners_2D'                  : n2,
        'n_corners_3D_REG_bilinear'     : n_rb_bil,
        'n_corners_3D_REG_bicubic'      : n_rb_bic,
        'n_corners_3D_PC_bilinear'      : n_pc_bil,
        'n_corners_3D_PC_bicubic'       : n_pc_bic,
        'elapsed_s'                     : round(elapsed, 2),
        'status'                        : 'ok',
    }


def _empty_row(res_reg: float, res_pc: float, status: str) -> dict:
    nan = float('nan')
    return {
        'res_reg_mm_pix'                : res_reg,
        'res_pc_mm_pix'                 : res_pc,
        '2D_mean_px'                    : nan, '2D_median_px'                  : nan,
        '2D_mean_mm'                    : nan, '2D_median_mm'                  : nan,
        '3D_REG_bilinear_mean_mm'       : nan, '3D_REG_bilinear_median_mm'     : nan,
        '3D_REG_bicubic_mean_mm'        : nan, '3D_REG_bicubic_median_mm'      : nan,
        '3D_PC_bilinear_mean_mm'        : nan, '3D_PC_bilinear_median_mm'      : nan,
        '3D_PC_bicubic_mean_mm'         : nan, '3D_PC_bicubic_median_mm'       : nan,
        'n_corners_2D'                  : 0,
        'n_corners_3D_REG_bilinear'     : 0, 'n_corners_3D_REG_bicubic'      : 0,
        'n_corners_3D_PC_bilinear'      : 0, 'n_corners_3D_PC_bicubic'       : 0,
        'elapsed_s'                     : 0.0,
        'status'                        : status,
    }


# =============================================================================
# Selezione coppia migliore
# =============================================================================

def _find_best_row(rows: list[dict]) -> tuple[int, float]:
    """
    Trova l'indice della riga con il 3D error assoluto minimo.

    Per ogni riga prende il minimo tra le quattro colonne 3D mean
    (REG bilinear, REG bicubic, PC bilinear, PC bicubic), poi seleziona
    la riga con quel valore piu' basso.

    Returns
    -------
    (best_idx, best_value)   best_idx = -1 se nessuna riga valida.
    """
    best_idx = -1
    best_val = float('inf')

    for i, row in enumerate(rows):
        if row.get('status') != 'ok':
            continue
        vals  = [row.get(k, float('nan')) for k in _3D_ERROR_KEYS]
        valid = [v for v in vals if not np.isnan(v)]
        if not valid:
            continue
        row_min = min(valid)
        if row_min < best_val:
            best_val = row_min
            best_idx = i

    return best_idx, (best_val if best_idx >= 0 else float('nan'))


def _winner_key(row: dict) -> Optional[str]:
    """
    Ritorna la chiave della colonna 3D mean che ha il valore minimo nella riga.
    None se nessun valore valido.
    """
    best_k = None
    best_v = float('inf')
    for k in _3D_ERROR_KEYS:
        v = row.get(k, float('nan'))
        if not np.isnan(v) and v < best_v:
            best_v = v
            best_k = k
    return best_k


# =============================================================================
# Scrittura Excel riepilogo
# =============================================================================

def _write_summary_xlsx(
    rows: list[dict],
    output_path: str,
    sample_name: str,
    best_row_idx: int = -1,
) -> None:
    """
    Scrive l'Excel riepilogo dello sweep.

    La riga migliore (best_row_idx) viene evidenziata:
      - tutta la riga in verde (#00B050, testo bianco bold)
      - la cella con il valore minimo assoluto 3D in arancio (#FF8C00)

    In riga 3 viene riportato un riepilogo testuale della coppia migliore.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sweep Summary'

    # ── Stili base ───────────────────────────────────────────────────────────
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', start_color='305496')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin       = Side(border_style='thin', color='BFBFBF')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center')
    norm_font  = Font(name='Calibri', size=10)

    res_fill   = PatternFill('solid', start_color='FFF2CC')
    err2d_fill = PatternFill('solid', start_color='DDEBF7')
    reg_fill   = PatternFill('solid', start_color='D9E1F2')
    pc_fill    = PatternFill('solid', start_color='E2EFDA')
    meta_fill  = PatternFill('solid', start_color='F2F2F2')

    # ── Stili riga migliore ──────────────────────────────────────────────────
    best_fill   = PatternFill('solid', start_color='00B050')   # verde
    best_font   = Font(bold=True, color='FFFFFF', size=10)
    winner_fill = PatternFill('solid', start_color='FF8C00')   # arancio
    winner_font = Font(bold=True, color='FFFFFF', size=10)

    # ── Righe di intestazione ────────────────────────────────────────────────
    ws.cell(row=1, column=1,
            value=f"Resolution Sweep — {sample_name}").font = \
        Font(bold=True, size=14, color='1F4E78')
    ws.cell(row=2, column=1,
            value=f"Run: {len(rows)} configurazioni  "
                  f"|  Generato: {time.strftime('%Y-%m-%d %H:%M:%S')}").font = \
        Font(italic=True, color='595959')

    if best_row_idx >= 0:
        br  = rows[best_row_idx]
        bv  = _best_3d_val(br)
        bv_str = f"{bv:.4f}" if not np.isnan(bv) else "N/A"
        ws.cell(row=3, column=1,
                value=f"★ Best: res_reg={br['res_reg_mm_pix']} mm/px  "
                      f"res_pc={br['res_pc_mm_pix']} mm/px  "
                      f"→ 3D error min = {bv_str} mm  "
                      f"(riga {best_row_idx + 1})").font = \
            Font(italic=True, bold=True, color='00B050', size=11)

    # ── Definizione colonne ──────────────────────────────────────────────────
    columns: list[tuple[str, str, PatternFill]] = [
        ('res_reg_mm_pix',            'res_reg\n(mm/px)',              res_fill),
        ('res_pc_mm_pix',             'res_pc\n(mm/px)',               res_fill),
        ('2D_mean_px',                '2D mean\n(px)',                 err2d_fill),
        ('2D_median_px',              '2D median\n(px)',               err2d_fill),
        ('2D_mean_mm',                '2D mean\n(mm)',                 err2d_fill),
        ('2D_median_mm',              '2D median\n(mm)',               err2d_fill),
        ('3D_REG_bilinear_mean_mm',   '3D REG bilinear\nmean (mm)',    reg_fill),
        ('3D_REG_bilinear_median_mm', '3D REG bilinear\nmedian (mm)',  reg_fill),
        ('3D_REG_bicubic_mean_mm',    '3D REG bicubic\nmean (mm)',     reg_fill),
        ('3D_REG_bicubic_median_mm',  '3D REG bicubic\nmedian (mm)',   reg_fill),
        ('3D_PC_bilinear_mean_mm',    '3D PC bilinear\nmean (mm)',     pc_fill),
        ('3D_PC_bilinear_median_mm',  '3D PC bilinear\nmedian (mm)',   pc_fill),
        ('3D_PC_bicubic_mean_mm',     '3D PC bicubic\nmean (mm)',      pc_fill),
        ('3D_PC_bicubic_median_mm',   '3D PC bicubic\nmedian (mm)',    pc_fill),
        ('n_corners_2D',              'N corners\n2D',                 meta_fill),
        ('n_corners_3D_REG_bilinear', 'N corners\n3D REG bil',         meta_fill),
        ('n_corners_3D_REG_bicubic',  'N corners\n3D REG bic',         meta_fill),
        ('n_corners_3D_PC_bilinear',  'N corners\n3D PC bil',          meta_fill),
        ('n_corners_3D_PC_bicubic',   'N corners\n3D PC bic',          meta_fill),
        ('elapsed_s',                 'Elapsed\n(s)',                  meta_fill),
        ('status',                    'Status',                        meta_fill),
    ]

    HEADER_ROW = 5   # riga 4 libera per spaziatura

    for ci, (_, label, _) in enumerate(columns, start=1):
        c = ws.cell(row=HEADER_ROW, column=ci, value=label)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = border
    ws.row_dimensions[HEADER_ROW].height = 36

    # Determina la chiave della cella vincitrice (arancio)
    w_key = _winner_key(rows[best_row_idx]) if best_row_idx >= 0 else None

    def _fmt(v):
        return 'N/A' if isinstance(v, float) and np.isnan(v) else v

    for row_i, row in enumerate(rows, start=HEADER_ROW + 1):
        is_best = (row_i - HEADER_ROW - 1) == best_row_idx

        for ci, (key, _, fill) in enumerate(columns, start=1):
            val = row.get(key, '')
            c   = ws.cell(row=row_i, column=ci, value=_fmt(val))
            c.alignment = center
            c.border    = border

            if is_best:
                if key == w_key:
                    c.font = winner_font
                    c.fill = winner_fill
                else:
                    c.font = best_font
                    c.fill = best_fill
            else:
                c.font = norm_font
                c.fill = fill

            if isinstance(val, float) and not np.isnan(val):
                if key.endswith(('_mm', '_px')) or key == 'elapsed_s':
                    c.number_format = '0.0000'

    for ci, (_, label, _) in enumerate(columns, start=1):
        max_len = max(len(line) for line in label.split('\n'))
        ws.column_dimensions[get_column_letter(ci)].width = max(max_len + 3, 12)

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"\n[Sweep] Excel riepilogo salvato -> {output_path}")


# =============================================================================
# Entry point dello sweep
# =============================================================================

def run_sweep(cfg: dict, aruco_dict_cv: int, torch_device: Optional[str],
              use_torch_render: bool, render_fn=None) -> str:
    """
    Esegue lo sweep completo e ritorna il path dell'Excel riepilogo.

    Se cfg['export']['save_pointcloud'] o cfg['export']['save_images'] sono True,
    al termine viene rieseguita la coppia migliore (3D error minimo assoluto)
    con i flag di salvataggio attivi. L'output e' scritto in output_dir.
    """
    sweep_cfg = cfg.get('sweep', {}) or {}
    pairs = parse_pairs(sweep_cfg)

    save_pointcloud_cfg = cfg['export'].get('save_pointcloud', False)
    save_images_cfg     = cfg['export'].get('save_images', False)
    do_best_rerun       = save_pointcloud_cfg or save_images_cfg

    print("\n" + "=" * 68)
    print(f"  SWEEP RESOLUTION  —  {len(pairs)} coppie da testare")
    print("=" * 68)
    for i, (r, p) in enumerate(pairs):
        print(f"  [{i + 1:2d}/{len(pairs)}]  res_reg = {r:>6.3f} mm/px   "
              f"res_pc = {p:>6.3f} mm/px")
    if do_best_rerun:
        print(f"\n  [INFO] Al termine verra' rieseguita la coppia migliore con "
              f"save_pointcloud={save_pointcloud_cfg}  "
              f"save_images={save_images_cfg}")
    print()

    base_out = cfg['paths']['output_dir']
    os.makedirs(base_out, exist_ok=True)
    sweep_tmp_root = os.path.join(base_out, '_sweep_tmp')
    os.makedirs(sweep_tmp_root, exist_ok=True)
    print(f"[Sweep] Subdir temporanee in: {sweep_tmp_root}")
    print(f"[Sweep] (verranno eliminate al termine dello sweep)\n")

    # ── Pre-calcolo render unici ─────────────────────────────────────────────
    render_cache: Optional[dict[float, tuple]] = None
    mesh = None

    if use_torch_render and render_fn is not None:
        print(f"[Sweep] Carico mesh: {cfg['paths']['mesh']}")
        mesh = load_mesh(cfg['paths']['mesh'], scale_m_to_mm=True)

        unique_res = _collect_unique_resolutions(pairs)
        print(f"\n[Sweep] Risoluzioni uniche: {[str(r) for r in unique_res]}")
        print(f"[Sweep] -> {len(unique_res)} render invece di "
              f"{2*len(pairs)} potenziali "
              f"({2*len(pairs) - len(unique_res)} evitati)")

        render_cache = _precompute_renders(
            mesh=mesh,
            unique_resolutions=unique_res,
            margin_mm=cfg['render']['margin_mm'],
            torch_device=torch_device,
            render_fn=render_fn,
        )
    else:
        print("[Sweep] AVVISO: render_fn non disponibile -> cache disabilitata.")
        print("[Sweep] Ogni run rifara' il render internamente (fallback CPU).")

    # ── Loop principale (nessun salvataggio) ─────────────────────────────────
    rows: list[dict] = []
    for i, (res_reg, res_pc) in enumerate(pairs):
        tag     = f"reg{str(res_reg).replace('.','p')}_pc{str(res_pc).replace('.','p')}"
        tmp_dir = os.path.join(sweep_tmp_root, f"run{i+1:02d}_{tag}")

        try:
            row = _run_single_pair(
                cfg=cfg,
                aruco_dict_cv=aruco_dict_cv,
                res_reg=res_reg,
                res_pc=res_pc,
                torch_device=torch_device,
                use_torch_render=use_torch_render,
                tmp_dir=tmp_dir,
                render_cache=render_cache,
                save_pointcloud_override=False,
                save_images_override=False,
            )
        except Exception as e:
            print(f"\n[Sweep] ERRORE su coppia ({res_reg}, {res_pc}): {e}")
            traceback.print_exc()
            row = _empty_row(res_reg, res_pc, status=f'error: {type(e).__name__}: {e}')

        rows.append(row)
        print(f"\n[Sweep] Progresso: {i + 1}/{len(pairs)} run completate.\n")

    # ── Trova la coppia migliore ─────────────────────────────────────────────
    best_idx, best_val = _find_best_row(rows)

    if best_idx >= 0:
        br = rows[best_idx]
        print(f"\n[Sweep] ★ Coppia migliore (riga {best_idx+1}/{len(rows)}): "
              f"res_reg={br['res_reg_mm_pix']}  res_pc={br['res_pc_mm_pix']}  "
              f"-> 3D error min = {best_val:.4f} mm")
    else:
        print("\n[Sweep] AVVISO: nessuna riga valida trovata per il best.")

    # ── Re-run della coppia migliore con salvataggio (se richiesto) ──────────
    if do_best_rerun and best_idx >= 0:
        br           = rows[best_idx]
        res_reg_best = br['res_reg_mm_pix']
        res_pc_best  = br['res_pc_mm_pix']

        print(f"\n[Sweep] Re-run coppia migliore con salvataggio...")
        print(f"  res_reg={res_reg_best}  res_pc={res_pc_best}")
        print(f"  save_pointcloud={save_pointcloud_cfg}  "
              f"save_images={save_images_cfg}")
        print(f"  Output -> {base_out}")

        try:
            _run_single_pair(
                cfg=cfg,
                aruco_dict_cv=aruco_dict_cv,
                res_reg=res_reg_best,
                res_pc=res_pc_best,
                torch_device=torch_device,
                use_torch_render=use_torch_render,
                tmp_dir=base_out,          # direttamente nella output_dir finale
                render_cache=render_cache,
                save_pointcloud_override=save_pointcloud_cfg,
                save_images_override=save_images_cfg,
            )
            print(f"[Sweep] Re-run completato. Output in: {base_out}")
        except Exception as e:
            print(f"[Sweep] ERRORE nel re-run della coppia migliore: {e}")
            traceback.print_exc()

    # ── Pulizia ──────────────────────────────────────────────────────────────
    try:
        shutil.rmtree(sweep_tmp_root)
        print(f"[Sweep] Cartella temporanea rimossa: {sweep_tmp_root}")
    except Exception as e:
        print(f"[Sweep] Avviso: impossibile rimuovere {sweep_tmp_root}: {e}")

    # ── Excel riepilogo ──────────────────────────────────────────────────────
    sample   = cfg['sample_name']
    out_path = os.path.join(base_out, f"{sample}_sweep_resolution_summary.xlsx")
    _write_summary_xlsx(rows, out_path, sample_name=sample, best_row_idx=best_idx)

    # ── Riepilogo finale ─────────────────────────────────────────────────────
    print("\n" + "=" * 68)
    print("SWEEP COMPLETATO")
    print("=" * 68)
    print(f"  Coppie eseguite : {len(rows)}")
    n_ok = sum(1 for r in rows if r['status'] == 'ok')
    print(f"  OK / Falliti    : {n_ok} / {len(rows) - n_ok}")
    if best_idx >= 0:
        br = rows[best_idx]
        print(f"  ★ Coppia migliore: res_reg={br['res_reg_mm_pix']}  "
              f"res_pc={br['res_pc_mm_pix']}  3D error={best_val:.4f} mm")
        if do_best_rerun:
            print(f"  Output best     : {base_out}")
    print(f"  Output Excel    : {out_path}")
    print("=" * 68)

    return out_path
