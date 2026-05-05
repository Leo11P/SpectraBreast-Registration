"""
HSI -> MESH 3D REGISTRATION  +  SPECTRAL POINT CLOUD BUILDER
=============================================================
Single autonomous file — no external pipeline imports required.

What it does (in order):
  1.  Load ArUco 3D corners from JSON  (metres -> mm)
  2.  Load HSI cube, extract 2D image for ArUco detection
  3.  Load mesh, scale m->mm
  4.  Build orthographic top-view render via chunked vertical ray casting
        -> render_rgb, depth_map, xyz_map
  5.  Detect ArUco on HSI 2D image
  6.  Project JSON 3D corners onto render pixel space  (analytical, no intrinsics)
  7.  Match HSI corners <-> render corners, compute homography H (RANSAC)
  8.  Project suspicious HSI points -> render pixel -> (X,Y,Z) on mesh
  9.  Compute quality metrics:
        - 2D reprojection error on ArUco corners  (px  and  mm)
        - 3D reprojection error vs JSON ground truth  (mm)
          calcolato con DUE metodi di interpolazione:
            * bilineare  (4 vicini, 2x2)
            * bicubico   (Catmull-Rom 4x4 con fallback bilineare ai bordi)
          La pipeline usa automaticamente il metodo con errore medio minore
          per i suspicious points e per il logging finale.
  10. Save Excel report con entrambe le metriche 3D, statistiche per-marker
      (mean + median) e righe finali di mean+median globali.
  11. Build spectral point cloud:
        for every render pixel with valid (X,Y,Z):
          back-project via H_inv -> HSI pixel -> read full spectrum
  12. Export point cloud:  .ply  (CloudCompare),  .npz  (Python),  .csv  (opt.)
  13. Save coverage visualisation + TURBO depth render

Dependencies:
    pip install trimesh rtree opencv-contrib-python numpy openpyxl
    (optional) pip install pyembree   -- for faster ray casting

Coordinate conventions:
    JSON / mesh  : metres input, converted to mm internally
    xyz_map      : mm, same axes as mesh after apply_scale(1000)
    Render pixel : col = X direction, row = Y direction (top-left origin)
    HSI pixel    : col = sample axis, row = line axis
"""

import cv2
import json
import numpy as np
import os
import trimesh
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# 0 — HSI loading (ENVI format)
# =============================================================================

def load_envi(hdr_path):
    """
    Load an ENVI hyperspectral cube from a .hdr file.

    Returns
    -------
    cube : (rows, cols, bands)  float32
    meta : dict  parsed header key/value pairs
    """
    meta = {}
    with open(hdr_path, 'r') as f:
        for line in f:
            line = line.strip()
            if '=' in line:
                key, val = line.split('=', 1)
                meta[key.strip().lower()] = val.strip()

    rows   = int(meta['lines'])
    cols   = int(meta['samples'])
    bands  = int(meta['bands'])
    offset = int(meta.get('header offset', 0))

    dtype_code = int(meta.get('data type', 4))
    interleave = meta.get('interleave', 'bsq').lower()

    dtype_map = {1: np.uint8, 2: np.int16, 4: np.float32, 5: np.float64}
    dtype = dtype_map.get(dtype_code, np.float32)

    base = os.path.splitext(hdr_path)[0]
    data_path = None
    for ext in ['', '.raw', '.bil', '.bip', '.bsq', '.img']:
        candidate = base + ext
        if os.path.exists(candidate) and candidate != hdr_path:
            data_path = candidate
            break

    if not data_path:
        raise FileNotFoundError(f"Data file not found for: {hdr_path}")

    with open(data_path, 'rb') as f:
        f.seek(offset)
        raw = np.fromfile(f, dtype=dtype)

    expected_size = rows * cols * bands
    if raw.size != expected_size:
        print(f"[Warning] Expected {expected_size} values, found {raw.size}. Clipping...")
        raw = raw[:expected_size]

    if interleave == 'bsq':
        cube = raw.reshape((bands, rows, cols)).transpose(1, 2, 0)
    elif interleave == 'bil':
        cube = raw.reshape((rows, bands, cols)).transpose(0, 2, 1)
    elif interleave == 'bip':
        cube = raw.reshape((rows, cols, bands))
    else:
        raise ValueError(f"Unsupported interleave: {interleave}")

    print(f"[HSI] Loaded: {rows}x{cols}x{bands} ({interleave.upper()}), offset={offset}")
    return cube, meta


def extract_2d_from_hsi(cube, meta, method='visible_band'):
    """
    Extract a single 2D grayscale image from an HSI cube for ArUco detection.

    Parameters
    ----------
    method : 'visible_band' | 'mean' | 'pca'
    """
    if method == 'visible_band':
        wavelengths_str = meta.get('wavelength', None)
        if wavelengths_str:
            wavelengths_str = wavelengths_str.strip('{}').strip()
            wavelengths = np.array([float(w) for w in wavelengths_str.split(',')])
            band_idx = np.argmin(np.abs(wavelengths - 550.0))
            print(f"[HSI] Selected band: idx={band_idx}, lambda={wavelengths[band_idx]:.1f}nm")
        else:
            band_idx = cube.shape[2] // 2
            print(f"[HSI] No wavelength info — using central band idx={band_idx}")
        img_2d = cube[:, :, band_idx].astype(np.float32)

    elif method == 'pca':
        rows, cols, bands = cube.shape
        X = cube.reshape(-1, bands).astype(np.float32)
        X -= X.mean(axis=0)
        _, _, Vt = np.linalg.svd(X, full_matrices=False)
        pc1 = X @ Vt[0]
        img_2d = pc1.reshape(rows, cols)
        print(f"[HSI] PCA: first component extracted")

    elif method == 'mean':
        img_2d = cube.mean(axis=2).astype(np.float32)
        print(f"[HSI] Mean of {cube.shape[2]} bands")

    else:
        raise ValueError(f"Invalid method: {method}")

    img_2d = np.nan_to_num(img_2d)
    img_2d = np.maximum(img_2d, 0)

    p2, p98 = np.percentile(img_2d, (2, 98))
    if p98 - p2 > 0:
        img_norm = np.clip((img_2d - p2) / (p98 - p2), 0, 1)
    else:
        img_norm = img_2d / (np.max(img_2d) + 1e-6)

    img_8bit = (img_norm * 255).astype(np.uint8)
    img_8bit = cv2.equalizeHist(img_8bit)
    return img_8bit


# =============================================================================
# 0b — ArUco detection
# =============================================================================

def detect_aruco(image_gray, aruco_dict_type=cv2.aruco.DICT_4X4_50,
                 use_subpix=True, subpix_winsize=5,
                 subpix_maxiter=30, subpix_eps=0.001):
    """
    Detect ArUco markers in a grayscale image.

    Dopo la detection standard, affina le posizioni dei corner a livello
    sub-pixel con cv2.cornerSubPix (Punto 3). Questo riduce l'errore di
    localizzazione da ±0.5-1 px a ±0.05-0.2 px, producendo un'omografia
    più accurata e un errore 3D più basso.

    Parameters
    ----------
    image_gray      : (H, W) uint8
    aruco_dict_type : dizionario ArUco OpenCV
    use_subpix      : abilita il refinement sub-pixel
    subpix_winsize  : semi-finestra per cornerSubPix (default 5 → finestra 11×11)
    subpix_maxiter  : iterazioni massime per cornerSubPix
    subpix_eps      : epsilon di convergenza per cornerSubPix

    Returns
    -------
    found_data : dict { marker_id (int): corners (4,2) float32 }
    ids        : raw ids array da OpenCV (o None)
    """
    aruco_dict   = cv2.aruco.getPredefinedDictionary(aruco_dict_type)
    aruco_params = cv2.aruco.DetectorParameters()
    detector     = cv2.aruco.ArucoDetector(aruco_dict, aruco_params)

    corners, ids, _ = detector.detectMarkers(image_gray)

    if ids is None:
        return {}, None

    if use_subpix:
        criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER,
                    subpix_maxiter, subpix_eps)
        win  = (subpix_winsize, subpix_winsize)
        dead = (-1, -1)

        corners_before = np.vstack([c[0] for c in corners])   # (N*4, 2)
        corners = [
            cv2.cornerSubPix(
                image_gray, c[0].copy(), win, dead, criteria
            )[np.newaxis]                                      # (1, 4, 2)
            for c in corners
        ]
        corners_after = np.vstack([c[0] for c in corners])
        delta = np.linalg.norm(corners_after - corners_before, axis=1)
        print(f"  [SubPix] shift medio={delta.mean():.4f} px  "
              f"max={delta.max():.4f} px  "
              f"(finestra {subpix_winsize*2+1}×{subpix_winsize*2+1})")

    found_data = {int(ids[i][0]): corners[i][0] for i in range(len(ids))}
    return found_data, ids


# =============================================================================
# 1 — ArUco 3D from JSON
# =============================================================================

def load_aruco_3d_json(json_path, scale_m_to_mm=True):
    """
    Load ArUco 3D corner positions from a JSON file.

    Expected JSON structure:
        { "markers": { "0": { "corners_3d": [[x,y,z], ...] }, ... } }

    Returns
    -------
    aruco_3d : dict { marker_id (int): corners (4,3) float32 }  in mm
    meta     : full parsed JSON dict
    """
    with open(json_path, 'r') as f:
        meta = json.load(f)
    scale    = 1000.0 if scale_m_to_mm else 1.0
    aruco_3d = {}
    for id_str, mdata in meta['markers'].items():
        marker_id           = int(id_str)
        corners             = np.array(mdata['corners_3d'], dtype=np.float32) * scale
        aruco_3d[marker_id] = corners
    unit = 'mm' if scale_m_to_mm else 'm'
    print(f"[ArUco3D] Loaded {len(aruco_3d)} markers ({unit}): {sorted(aruco_3d.keys())}")
    for mid, c in aruco_3d.items():
        ctr  = c.mean(axis=0)
        side = float(np.mean([np.linalg.norm(c[(j+1)%4]-c[j]) for j in range(4)]))
        print(f"  ID={mid}: center=({ctr[0]:.1f},{ctr[1]:.1f},{ctr[2]:.1f}) mm  "
              f"mean_side={side:.2f} mm")
    return aruco_3d, meta


# =============================================================================
# 2 — Load mesh
# =============================================================================

def load_mesh(mesh_path, scale_m_to_mm=True):
    """
    Load a 3D mesh and optionally scale from metres to millimetres.

    Returns
    -------
    mesh : trimesh.Trimesh
    """
    mesh = trimesh.load(mesh_path, force='mesh')
    if not isinstance(mesh, trimesh.Trimesh):
        mesh = trimesh.util.concatenate(
            list(mesh.geometry.values()) if hasattr(mesh, 'geometry') else [mesh]
        )
    print(f"[Mesh] Loaded : {mesh_path}")
    print(f"       Vertices: {len(mesh.vertices)}   Faces: {len(mesh.faces)}")
    print(f"       Bounds  : {mesh.bounds}")
    if scale_m_to_mm:
        mesh.apply_scale(1000.0)
        print(f"[Mesh] Scaled x1000 (m -> mm). New bounds: {mesh.bounds}")
    return mesh


# =============================================================================
# 3 — Orthographic render with chunked ray casting
# =============================================================================

def _get_ray_intersector(mesh):
    try:
        from trimesh.ray.ray_pyembree import RayMeshIntersector
        intersector = RayMeshIntersector(mesh)
        print("[Render] Backend: pyembree (fast)")
        return intersector
    except Exception:
        print("[Render] Backend: trimesh BVH (pyembree not available)")
        return mesh.ray


def render_orthographic_topview(mesh, resolution_mm_per_px=2.0, margin_mm=10.0,
                                 chunk_size=500_000,
                                 xmin_override=None, xmax_override=None,
                                 ymin_override=None, ymax_override=None):
    """
    Build an orthographic top-view render of the mesh via chunked vertical
    ray casting.

    Optional *_override parameters allow rendering a sub-region (fine crop).

    Returns
    -------
    render_rgb  : (H, W, 3) uint8    depth-shaded grayscale as BGR
    depth_map   : (H, W)    float32  Z in mm, NaN = no mesh hit
    xyz_map     : (H, W, 3) float32  (X,Y,Z) in mm per pixel
    origin_xy   : (xmin, ymin) top-left corner in mesh mm coordinates
    res         : resolution_mm_per_px (echoed)
    """
    bounds = mesh.bounds
    xmin = xmin_override if xmin_override is not None else bounds[0, 0] - margin_mm
    xmax = xmax_override if xmax_override is not None else bounds[1, 0] + margin_mm
    ymin = ymin_override if ymin_override is not None else bounds[0, 1] - margin_mm
    ymax = ymax_override if ymax_override is not None else bounds[1, 1] + margin_mm
    zmax = float(bounds[1, 2]) + 10.0

    xs = np.arange(xmin, xmax, resolution_mm_per_px)
    ys = np.arange(ymin, ymax, resolution_mm_per_px)
    W, H = len(xs), len(ys)

    print(f"[Render] Grid   : {W} x {H} px  ({resolution_mm_per_px} mm/px)")
    print(f"         X: [{xmin:.1f}, {xmax:.1f}] mm   Y: [{ymin:.1f}, {ymax:.1f}] mm")

    xx, yy      = np.meshgrid(xs, ys)
    n_rays      = xx.size
    ray_ori_all = np.column_stack([xx.ravel(), yy.ravel(),
                                   np.full(n_rays, zmax)]).astype(np.float64)
    del xx, yy
    ray_dir_single = np.array([[0.0, 0.0, -1.0]])

    print(f"[Render] Casting {n_rays:,} rays in chunks of {chunk_size:,} ...")

    depth_map = np.full((H, W), np.nan, dtype=np.float32)
    xyz_map   = np.full((H, W, 3), np.nan, dtype=np.float32)

    intersector = _get_ray_intersector(mesh)
    total_hits  = 0
    n_chunks    = (n_rays + chunk_size - 1) // chunk_size

    for i in range(n_chunks):
        start     = i * chunk_size
        end       = min(start + chunk_size, n_rays)
        chunk_ori = ray_ori_all[start:end]
        chunk_dir = np.tile(ray_dir_single, (len(chunk_ori), 1))

        locs, idx_ray, _ = intersector.intersects_location(
            ray_origins=chunk_ori, ray_directions=chunk_dir, multiple_hits=False
        )
        if len(locs) > 0:
            global_idx = start + idx_ray
            r_hit, c_hit = global_idx // W, global_idx % W
            depth_map[r_hit, c_hit]    = locs[:, 2].astype(np.float32)
            xyz_map  [r_hit, c_hit, 0] = locs[:, 0].astype(np.float32)
            xyz_map  [r_hit, c_hit, 1] = locs[:, 1].astype(np.float32)
            xyz_map  [r_hit, c_hit, 2] = locs[:, 2].astype(np.float32)
            total_hits += len(locs)
        print(f"  Chunk {i+1}/{n_chunks}  hits so far: {total_hits:,}", end='\r')

    del ray_ori_all
    print(f"\n[Render] Hits: {total_hits:,} / {n_rays:,}")

    z_valid = depth_map[~np.isnan(depth_map)]
    if z_valid.size > 0:
        z_min, z_max = z_valid.min(), z_valid.max()
        denom = (z_max - z_min) if z_max > z_min else 1.0
        gray  = np.where(np.isnan(depth_map), 0,
                         (depth_map - z_min) / denom * 255).astype(np.uint8)
    else:
        gray = np.zeros((H, W), dtype=np.uint8)

    render_rgb = cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)
    origin_xy  = (xmin, ymin)
    print(f"[Render] Done. origin_xy = ({xmin:.2f}, {ymin:.2f}) mm")
    return render_rgb, depth_map, xyz_map, origin_xy, resolution_mm_per_px


def save_render(render_rgb, depth_map, output_dir='.', prefix='render'):
    os.makedirs(output_dir, exist_ok=True)
    rp = os.path.join(output_dir, f'{prefix}_topview.png')
    dp = os.path.join(output_dir, f'{prefix}_depth.npy')
    cv2.imwrite(rp, render_rgb)
    if depth_map is not None:
        np.save(dp, depth_map)
    print(f"[Render] Saved: {rp}")
    return rp


# =============================================================================
# 4 — ArUco: project JSON corners -> render pixel space
# =============================================================================

def project_aruco3d_to_render(aruco_3d, origin_xy, resolution_mm_per_px):
    """
    Convert 3D ArUco corners (mm) to render pixel coordinates via orthographic
    projection:
        col = (X_mm - xmin) / resolution_mm_per_px
        row = (Y_mm - ymin) / resolution_mm_per_px

    Returns
    -------
    data_render : dict { marker_id: corners_px (4,2) float32 }
    """
    xmin, ymin  = origin_xy
    data_render = {}
    for mid, corners_3d in aruco_3d.items():
        cols = (corners_3d[:, 0] - xmin) / resolution_mm_per_px
        rows = (corners_3d[:, 1] - ymin) / resolution_mm_per_px
        data_render[mid] = np.column_stack([cols, rows]).astype(np.float32)
    print(f"[Project3D->render] {len(data_render)} markers projected.")
    return data_render


# =============================================================================
# 5 — Match & homography
# =============================================================================

def match_aruco_hsi_render(data_hsi, data_render, min_markers=1):
    """
    Match ArUco corners detected on HSI with those projected onto the render.

    Returns
    -------
    pts_hsi    : (N, 2) float32  HSI pixel coordinates
    pts_render : (N, 2) float32  render pixel coordinates
    common_ids : list of int     matched marker IDs
    """
    common_ids = set(data_hsi.keys()) & set(data_render.keys())
    if len(common_ids) < min_markers:
        raise ValueError(
            f"[match] Only {len(common_ids)} common markers (need >= {min_markers}). "
            f"HSI: {set(data_hsi.keys())}  JSON: {set(data_render.keys())}"
        )
    if len(common_ids) < 4:
        print(f"[match] WARNING: only {len(common_ids)} markers — "
              f"proceeding with {len(common_ids)*4} corner pairs.")
    pts_hsi, pts_render = [], []
    for mid in sorted(common_ids):
        pts_hsi.extend(data_hsi[mid])
        pts_render.extend(data_render[mid])
    print(f"[match] Common: {sorted(common_ids)}  ({len(common_ids)*4} pairs)")
    return (np.array(pts_hsi,    dtype=np.float32),
            np.array(pts_render, dtype=np.float32),
            list(common_ids))


def compute_homography(pts_src, pts_dst, tag=''):
    """
    Compute a homography from pts_src to pts_dst using RANSAC.

    Returns
    -------
    H    : (3,3) float64
    mask : inlier mask from RANSAC
    """
    H, mask = cv2.findHomography(pts_src, pts_dst, cv2.RANSAC,
                                  ransacReprojThreshold=3.0)
    if H is None:
        raise RuntimeError(f"Homography failed {tag}")
    inliers = int(mask.ravel().sum())
    print(f"[Homography{tag}] Inliers: {inliers}/{len(pts_src)}")
    print(f"[Homography{tag}] H =\n{H}")
    return H, mask


# =============================================================================
# 6 — 3D lookup (bilinear)
# =============================================================================

def _bilinear_lookup(pts_render_px, xyz_map):
    """
    Vectorised bilinear lookup of (X,Y,Z) from xyz_map at sub-pixel coords.
    For ogni punto: 4 vicini (2x2). Se almeno 1 valido -> media pesata
    rinormalizzata. Se nessuno valido -> NaN.
    """
    H, W = xyz_map.shape[:2]
    pts  = np.asarray(pts_render_px, dtype=np.float32)
    N    = pts.shape[0]
    if N == 0:
        return np.empty((0, 3), dtype=np.float32)

    col, row = pts[:, 0], pts[:, 1]
    c0 = np.floor(col).astype(np.int64); r0 = np.floor(row).astype(np.int64)
    c1, r1 = c0 + 1, r0 + 1
    dc, dr = (col - c0).astype(np.float32), (row - r0).astype(np.float32)

    in_bb = (c0 >= 0) & (r0 >= 0) & (c1 < W) & (r1 < H)
    c0c, c1c = np.clip(c0, 0, W-1), np.clip(c1, 0, W-1)
    r0c, r1c = np.clip(r0, 0, H-1), np.clip(r1, 0, H-1)

    v00 = xyz_map[r0c, c0c];  v10 = xyz_map[r0c, c1c]
    v01 = xyz_map[r1c, c0c];  v11 = xyz_map[r1c, c1c]
    m00 = ~np.isnan(v00[:,0]) & in_bb;  m10 = ~np.isnan(v10[:,0]) & in_bb
    m01 = ~np.isnan(v01[:,0]) & in_bb;  m11 = ~np.isnan(v11[:,0]) & in_bb

    w00 = ((1-dc)*(1-dr)) * m00;  w10 = (dc*(1-dr)) * m10
    w01 = ((1-dc)*dr)     * m01;  w11 = (dc*dr)     * m11
    v00 = np.where(m00[:,None], v00, 0.0); v10 = np.where(m10[:,None], v10, 0.0)
    v01 = np.where(m01[:,None], v01, 0.0); v11 = np.where(m11[:,None], v11, 0.0)

    num = (w00[:,None]*v00 + w10[:,None]*v10 +
           w01[:,None]*v01 + w11[:,None]*v11)
    den = (w00 + w10 + w01 + w11)

    out = np.full((N, 3), np.nan, dtype=np.float32)
    ok  = den > 0
    out[ok] = (num[ok] / den[ok, None]).astype(np.float32)
    return out


def _catmull_rom_weights(t):
    """
    Pesi Catmull-Rom (kernel cubico) per un offset frazionario t in [0, 1).
    Restituisce (N, 4): pesi per i 4 vicini lungo un asse, posizioni -1, 0, 1, 2.
    """
    t  = t.astype(np.float32)
    t2 = t * t
    t3 = t2 * t
    w_m1 = -0.5*t3 +     t2 - 0.5*t
    w_0  =  1.5*t3 - 2.5*t2 + 1.0
    w_p1 = -1.5*t3 + 2.0*t2 + 0.5*t
    w_p2 =  0.5*t3 - 0.5*t2
    return np.stack([w_m1, w_0, w_p1, w_p2], axis=1)   # (N, 4)


def _bicubic_lookup(pts_render_px, xyz_map):
    """
    Lookup bicubico (Catmull-Rom 4x4) di (X,Y,Z). Vettorizzato.
    Richiede tutti i 16 vicini validi (non-NaN); altrove restituisce NaN
    e il chiamante può decidere se fare fallback.
    """
    H, W = xyz_map.shape[:2]
    pts  = np.asarray(pts_render_px, dtype=np.float32)
    N    = pts.shape[0]
    if N == 0:
        return np.empty((0, 3), dtype=np.float32)

    col, row = pts[:, 0], pts[:, 1]
    c0 = np.floor(col).astype(np.int64); r0 = np.floor(row).astype(np.int64)
    dc, dr = (col - c0).astype(np.float32), (row - r0).astype(np.float32)

    # Vincolo bordi: per il kernel 4x4 serve l'intervallo [-1, +2]
    in_bb = (c0 - 1 >= 0) & (r0 - 1 >= 0) & (c0 + 2 < W) & (r0 + 2 < H)

    # Per evitare crash su out-of-bounds, clippiamo (le righe in_bb=False
    # verranno comunque scartate alla fine).
    c_idx = np.clip(c0[:, None] + np.array([-1, 0, 1, 2])[None, :], 0, W-1)  # (N,4)
    r_idx = np.clip(r0[:, None] + np.array([-1, 0, 1, 2])[None, :], 0, H-1)  # (N,4)

    # Costruisci la matrice 4x4 di vicini per ogni punto: shape (N, 4, 4, 3)
    # rs (N,4,1), cs (N,1,4) -> broadcast (N,4,4)
    rs = r_idx[:, :, None]
    cs = c_idx[:, None, :]
    neigh = xyz_map[rs, cs]                          # (N, 4, 4, 3)

    # Validità: nessun NaN nel 4x4 + entro bounds
    valid_all = (~np.isnan(neigh[..., 0])).all(axis=(1, 2)) & in_bb  # (N,)

    # Pesi Catmull-Rom su ciascun asse
    wx = _catmull_rom_weights(dc)                    # (N, 4)
    wy = _catmull_rom_weights(dr)                    # (N, 4)

    # Convoluzione separabile: prima lungo X (asse cs=ultimo asse della 4x4),
    # poi lungo Y. neigh[:, ry, rx, :] -> wx pesa rx
    # Step 1: combina su X => (N, 4, 3)  righe (rs)
    row_interp = np.einsum('nrxc,nx->nrc', neigh, wx)
    # Step 2: combina su Y => (N, 3)
    out_xyz    = np.einsum('nrc,nr->nc', row_interp, wy)

    out = np.full((N, 3), np.nan, dtype=np.float32)
    out[valid_all] = out_xyz[valid_all].astype(np.float32)
    return out


def lookup_3d(pts_render_px, xyz_map, mode='bilinear', verbose=True):
    """
    Lookup 3D di (X,Y,Z) da xyz_map a coordinate render sub-pixel.

    Parameters
    ----------
    pts_render_px : (N, 2)  (col, row) float
    xyz_map       : (H, W, 3) float32
    mode          : 'bilinear' | 'bicubic' | 'both'
        - 'bilinear' : interpolazione bilineare (4 vicini)
        - 'bicubic'  : kernel 4x4 Catmull-Rom; fallback automatico a
                       bilineare nei punti dove non sono disponibili
                       tutti e 16 i vicini validi (bordi mesh)
        - 'both'     : ritorna una tupla (coords_bilinear, coords_bicubic)
    verbose       : se True stampa una riga di sommario per chiamata

    Returns
    -------
    coords_3d : (N, 3) float32  oppure tupla di due (N, 3) se mode='both'
    """
    pts = np.asarray(pts_render_px, dtype=np.float32)
    N   = pts.shape[0]

    coords_bil = _bilinear_lookup(pts, xyz_map)

    if mode == 'bilinear':
        if verbose:
            valid = int(np.sum(~np.isnan(coords_bil[:, 0])))
            print(f"  [Lookup3D/bilinear] {valid}/{N} punti validi.")
        return coords_bil

    coords_bic_raw = _bicubic_lookup(pts, xyz_map)
    # Fallback: dove la bicubica è NaN ma la bilineare no, copia la bilineare
    nan_bic   = np.isnan(coords_bic_raw[:, 0])
    have_bil  = ~np.isnan(coords_bil[:, 0])
    fb_mask   = nan_bic & have_bil
    coords_bic = coords_bic_raw.copy()
    coords_bic[fb_mask] = coords_bil[fb_mask]

    if verbose:
        v_bil = int(np.sum(~np.isnan(coords_bil[:, 0])))
        v_bic_pure = int(np.sum(~np.isnan(coords_bic_raw[:, 0])))
        v_bic_fb   = int(np.sum(~np.isnan(coords_bic[:, 0])))
        print(f"  [Lookup3D] bilinear={v_bil}/{N}  "
              f"bicubic_pure={v_bic_pure}/{N}  "
              f"bicubic+fallback={v_bic_fb}/{N}")

    if mode == 'bicubic':
        return coords_bic
    if mode == 'both':
        return coords_bil, coords_bic
    raise ValueError(f"mode deve essere 'bilinear' | 'bicubic' | 'both', non {mode!r}")


# =============================================================================
# 7 — Quality metrics
# =============================================================================

def reprojection_error_2d(pts_hsi, pts_render_gt, H, tag='Homography'):
    """
    Compute 2D reprojection error (px) by projecting HSI corners via H and
    comparing against render ground-truth corners.
    """
    pts_proj = cv2.perspectiveTransform(
        pts_hsi.reshape(-1, 1, 2).astype(np.float32), H
    ).reshape(-1, 2)
    errors = np.linalg.norm(pts_proj - pts_render_gt, axis=1)
    print(f"\n[2D Reprojection — {tag}]")
    print(f"  Mean: {errors.mean():.4f} px   Max: {errors.max():.4f} px   "
          f"Min: {errors.min():.4f} px")
    for i, e in enumerate(errors):
        print(f"  Corner {i}: {e:.4f} px")
    return errors


def reprojection_error_3d_json(aruco_3d, data_hsi, data_render, xyz_map, H,
                               tag='reg'):
    """
    3D accuracy: project ArUco HSI corners -> render via H -> lookup (X,Y,Z),
    then compare against JSON ground-truth 3D positions.

    Calcola il lookup sia in bilineare che in bicubico (Catmull-Rom 4x4 con
    fallback bilineare ai bordi mesh).

    Parameters
    ----------
    tag : str  etichetta usata nei print (es. 'reg' o 'pc')

    Returns
    -------
    out : dict con chiavi
        'bilinear'  : (N,) float32  errori in mm (NaN dove invalido)
        'bicubic'   : (N,) float32  errori in mm (NaN dove invalido)
        'common_ids': list[int]     marker IDs effettivamente valutati
        'corner_ids': (N,) int      marker ID per ogni corner (in ordine)
        'corner_j'  : (N,) int      indice 0..3 del corner dentro al marker
        oppure None se non ci sono marker comuni
    """
    common = sorted(set(data_hsi.keys()) & set(data_render.keys()) & set(aruco_3d.keys()))
    if not common:
        print(f"[3D Error/{tag}] No common markers.")
        return None

    pts_hsi_all = np.vstack([data_hsi[m] for m in common]).astype(np.float32)
    gt_3d       = np.vstack([aruco_3d[m] for m in common]).astype(np.float32)
    proj_h      = cv2.perspectiveTransform(pts_hsi_all.reshape(-1,1,2), H).reshape(-1,2)

    coords_bil, coords_bic = lookup_3d(proj_h, xyz_map, mode='both')

    def _err(coords):
        valid  = ~np.isnan(gt_3d[:,0]) & ~np.isnan(coords[:,0])
        errors = np.full(len(coords), np.nan, dtype=np.float32)
        if valid.any():
            errors[valid] = np.linalg.norm(coords[valid] - gt_3d[valid], axis=1)
        return errors, int(valid.sum())

    err_bil, n_bil = _err(coords_bil)
    err_bic, n_bic = _err(coords_bic)

    print(f"\n[3D Error — {tag.upper()}]")
    if n_bil > 0:
        eb = err_bil[~np.isnan(err_bil)]
        print(f"  Bilinear ({n_bil:>2d} valid): mean={eb.mean():.4f}  "
              f"median={np.median(eb):.4f}  max={eb.max():.4f}  min={eb.min():.4f} mm")
    if n_bic > 0:
        ec = err_bic[~np.isnan(err_bic)]
        print(f"  Bicubic  ({n_bic:>2d} valid): mean={ec.mean():.4f}  "
              f"median={np.median(ec):.4f}  max={ec.max():.4f}  min={ec.min():.4f} mm")

    corner_ids = np.repeat(np.array(common, dtype=np.int64), 4)
    corner_j   = np.tile(np.arange(4, dtype=np.int64), len(common))

    return {
        'bilinear'  : err_bil,
        'bicubic'   : err_bic,
        'common_ids': common,
        'corner_ids': corner_ids,
        'corner_j'  : corner_j,
    }


# =============================================================================
# 8 — Excel error report
# =============================================================================

def _px_per_mm_from_data(data_hsi, marker_side_mm):
    """Compute global px/mm scale from detected HSI ArUco marker side lengths."""
    sides = []
    for corners in data_hsi.values():
        c = corners.astype(np.float32)
        sides.append(float(np.mean([np.linalg.norm(c[(j+1)%4]-c[j]) for j in range(4)])))
    return float(np.mean(sides)) / marker_side_mm if sides else 1.0


def save_excel_errors(err2d_px, err3d_dict, data_hsi, common_ids,
                      marker_side_mm, output_path,
                      err3d_dict_pc=None, res_reg=None, res_pc=None):
    """
    Save an Excel workbook con due sheet.

      Sheet 1 "Per-Corner Errors":
        - Errori per ogni corner:
            2D px | 2D mm | 3D REG bilinear | 3D REG bicubic
                           | 3D PC  bilinear | 3D PC  bicubic  (se err3d_dict_pc)
        - Blocco "Per-Marker Stats" (mean + median per marker)
        - Riga finale di mean e median globali

      Sheet 2 "Summary":
        - Sezione REG grid: statistiche 2D + 3D bilineare/bicubico
        - Sezione PC grid  (se err3d_dict_pc): stesse statistiche
        - Sezione Δ(PC - REG): confronto diretto

    Parameters
    ----------
    err2d_px       : (N,) float32  errori 2D reprojection in pixel
    err3d_dict     : dict da reprojection_error_3d_json (griglia REG) o None
    data_hsi       : dict marker_id -> corners HSI (per scala px/mm)
    common_ids     : list di marker_id usati per le 2D
    err3d_dict_pc  : dict da reprojection_error_3d_json (griglia PC)  o None
    res_reg        : float  risoluzione griglia REG in mm/px (per etichette)
    res_pc         : float  risoluzione griglia PC  in mm/px (per etichette)
    """
    px_per_mm = _px_per_mm_from_data(data_hsi, marker_side_mm)
    has_pc    = err3d_dict_pc is not None

    reg_lbl = f" REG ({res_reg} mm/px)" if res_reg else " REG"
    pc_lbl  = f" PC  ({res_pc} mm/px)"  if res_pc  else " PC"

    # --- Estrazione array 3D REG ---------------------------------------------
    if err3d_dict is not None:
        err3d_bil = np.asarray(err3d_dict.get('bilinear'), dtype=np.float64)
        err3d_bic = np.asarray(err3d_dict.get('bicubic'),  dtype=np.float64)
    else:
        err3d_bil = np.full(len(err2d_px), np.nan, dtype=np.float64)
        err3d_bic = np.full(len(err2d_px), np.nan, dtype=np.float64)

    # --- Estrazione array 3D PC (griglia point cloud) -------------------------
    if has_pc:
        err3d_pc_bil = np.asarray(err3d_dict_pc.get('bilinear'), dtype=np.float64)
        err3d_pc_bic = np.asarray(err3d_dict_pc.get('bicubic'),  dtype=np.float64)
    else:
        err3d_pc_bil = np.full(len(err2d_px), np.nan, dtype=np.float64)
        err3d_pc_bic = np.full(len(err2d_px), np.nan, dtype=np.float64)

    # Etichette (marker_id, j) per riga
    ids_sorted = sorted(common_ids) if common_ids else []
    if ids_sorted:
        ids_arr = np.repeat(np.asarray(ids_sorted), 4)
        j_arr   = np.tile(np.arange(4), len(ids_sorted))
    else:
        ids_arr = np.array([], dtype=np.int64)
        j_arr   = np.array([], dtype=np.int64)

    n_corners    = len(err2d_px)
    err2d_mm_arr = err2d_px / px_per_mm if px_per_mm > 0 else np.full_like(err2d_px, np.nan)

    # ========================================================================
    # Workbook setup
    # ========================================================================
    wb = Workbook()
    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', start_color='1F4E79')
    sub_fill  = PatternFill('solid', start_color='8FAADC')
    alt_fill  = PatternFill('solid', start_color='D6E4F0')
    tot_fill  = PatternFill('solid', start_color='FFE699')
    cen       = Alignment(horizontal='center', vertical='center')
    bd_side   = Side(style='thin', color='AAAAAA')
    bd        = Border(left=bd_side, right=bd_side, top=bd_side, bottom=bd_side)
    norm_font = Font(name='Calibri', size=10)
    bold_font = Font(name='Calibri', size=10, bold=True)

    def _hdr(ws, row, col, val, fill=hdr_fill):
        c = ws.cell(row=row, column=col, value=val)
        c.font, c.fill, c.alignment, c.border = hdr_font, fill, cen, bd

    def _cell(ws, row, col, val, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = bold_font if bold else norm_font
        c.alignment, c.border = cen, bd
        if fill:
            c.fill = fill

    def _r(x, nd=4):
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return 'N/A'
        return round(float(x), nd)

    def _stat(arr, fn):
        a = np.asarray(arr, dtype=np.float64)
        a = a[~np.isnan(a)]
        return float(fn(a)) if a.size else float('nan')

    # ========================================================================
    # Sheet 1: Per-Corner Errors
    # ========================================================================
    ws1 = wb.active
    ws1.title = 'Per-Corner Errors'

    n_3d_cols = 4 if has_pc else 2   # colonne 3D: REG bil, REG bic [, PC bil, PC bic]
    n_cols    = 5 + n_3d_cols        # 5 fissi + colonne 3D

    # Colori per distinguere i due blocchi 3D
    reg_fill = PatternFill('solid', start_color='D9E1F2')   # blu chiaro
    pc_fill  = PatternFill('solid', start_color='E2EFDA')   # verde chiaro

    headers = [
        'Corner index', 'Marker ID', 'Corner in marker',
        '2D error (px)', '2D error (mm)',
        f'3D{reg_lbl} bilinear (mm)', f'3D{reg_lbl} bicubic (mm)',
    ]
    if has_pc:
        headers += [f'3D{pc_lbl} bilinear (mm)', f'3D{pc_lbl} bicubic (mm)']

    for ci, h in enumerate(headers, 1):
        fill = reg_fill if ci in (6, 7) else (pc_fill if ci in (8, 9) else hdr_fill)
        _hdr(ws1, 1, ci, h, fill=fill)
    ws1.row_dimensions[1].height = 26

    # ---- Per-corner rows ----------------------------------------------------
    row_cursor = 2
    for ri in range(n_corners):
        fill = alt_fill if (row_cursor % 2 == 0) else None
        mid = int(ids_arr[ri]) if ri < len(ids_arr) else 'N/A'
        j   = int(j_arr[ri])   if ri < len(j_arr)   else 'N/A'
        _cell(ws1, row_cursor, 1, ri,                    fill)
        _cell(ws1, row_cursor, 2, mid,                   fill)
        _cell(ws1, row_cursor, 3, j,                     fill)
        _cell(ws1, row_cursor, 4, _r(err2d_px[ri]),      fill)
        _cell(ws1, row_cursor, 5, _r(err2d_mm_arr[ri]),  fill)
        _cell(ws1, row_cursor, 6, _r(err3d_bil[ri]    if ri < len(err3d_bil)    else np.nan), fill)
        _cell(ws1, row_cursor, 7, _r(err3d_bic[ri]    if ri < len(err3d_bic)    else np.nan), fill)
        if has_pc:
            _cell(ws1, row_cursor, 8, _r(err3d_pc_bil[ri] if ri < len(err3d_pc_bil) else np.nan), fill)
            _cell(ws1, row_cursor, 9, _r(err3d_pc_bic[ri] if ri < len(err3d_pc_bic) else np.nan), fill)
        row_cursor += 1

    # ---- Per-marker block ---------------------------------------------------
    row_cursor += 1
    _hdr(ws1, row_cursor, 1, 'PER-MARKER STATISTICS', sub_fill)
    for c in range(2, n_cols + 1):
        _hdr(ws1, row_cursor, c, '', sub_fill)
    row_cursor += 1

    for mid in ids_sorted:
        mask = (ids_arr == mid)
        if not mask.any():
            continue
        e2d_px = err2d_px[mask]
        e2d_mm = err2d_mm_arr[mask]
        e3b    = err3d_bil[mask]
        e3c    = err3d_bic[mask]
        e3pb   = err3d_pc_bil[mask] if has_pc else np.array([])
        e3pc   = err3d_pc_bic[mask] if has_pc else np.array([])

        for stat_name, stat_fn in (('mean', np.mean), ('median', np.median)):
            _cell(ws1, row_cursor, 1, f'Marker {mid}', tot_fill, bold=True)
            _cell(ws1, row_cursor, 2, mid,             tot_fill, bold=True)
            _cell(ws1, row_cursor, 3, stat_name,       tot_fill, bold=True)
            _cell(ws1, row_cursor, 4, _r(_stat(e2d_px, stat_fn)), tot_fill)
            _cell(ws1, row_cursor, 5, _r(_stat(e2d_mm, stat_fn)), tot_fill)
            _cell(ws1, row_cursor, 6, _r(_stat(e3b,    stat_fn)), tot_fill)
            _cell(ws1, row_cursor, 7, _r(_stat(e3c,    stat_fn)), tot_fill)
            if has_pc:
                _cell(ws1, row_cursor, 8, _r(_stat(e3pb, stat_fn)), tot_fill)
                _cell(ws1, row_cursor, 9, _r(_stat(e3pc, stat_fn)), tot_fill)
            row_cursor += 1

    # ---- Overall mean / median ---------------------------------------------
    row_cursor += 1
    _hdr(ws1, row_cursor, 1, 'OVERALL', sub_fill)
    for c in range(2, n_cols + 1):
        _hdr(ws1, row_cursor, c, '', sub_fill)
    row_cursor += 1

    for stat_name, stat_fn in (('mean', np.mean), ('median', np.median)):
        _cell(ws1, row_cursor, 1, f'ALL CORNERS — {stat_name}', tot_fill, bold=True)
        _cell(ws1, row_cursor, 2, '',          tot_fill)
        _cell(ws1, row_cursor, 3, stat_name,   tot_fill, bold=True)
        _cell(ws1, row_cursor, 4, _r(_stat(err2d_px,      stat_fn)), tot_fill, bold=True)
        _cell(ws1, row_cursor, 5, _r(_stat(err2d_mm_arr,  stat_fn)), tot_fill, bold=True)
        _cell(ws1, row_cursor, 6, _r(_stat(err3d_bil,     stat_fn)), tot_fill, bold=True)
        _cell(ws1, row_cursor, 7, _r(_stat(err3d_bic,     stat_fn)), tot_fill, bold=True)
        if has_pc:
            _cell(ws1, row_cursor, 8, _r(_stat(err3d_pc_bil, stat_fn)), tot_fill, bold=True)
            _cell(ws1, row_cursor, 9, _r(_stat(err3d_pc_bic, stat_fn)), tot_fill, bold=True)
        row_cursor += 1

    col_widths = [16, 12, 18, 16, 16, 24, 24, 24, 24]
    for ci, w in zip(range(1, n_cols + 1), col_widths):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ========================================================================
    # Sheet 2: Summary   (colonne: Metric | REG grid | PC grid)
    # ========================================================================
    ws2 = wb.create_sheet('Summary')

    # Header riga 1
    _hdr(ws2, 1, 1, 'Metric')
    _hdr(ws2, 1, 2, f'REG grid{(" (" + str(res_reg) + " mm/px)") if res_reg else ""}',
         fill=PatternFill('solid', start_color='2E75B6'))
    if has_pc:
        _hdr(ws2, 1, 3, f'PC grid{(" (" + str(res_pc) + " mm/px)") if res_pc else ""}',
             fill=PatternFill('solid', start_color='375623'))

    yf   = PatternFill('solid', start_color='FFF2CC')
    rf   = PatternFill('solid', start_color='D9E1F2')
    pf   = PatternFill('solid', start_color='E2EFDA')
    df   = PatternFill('solid', start_color='FCE4D6')   # arancione per Δ

    def _sum_row(ws, ri, label, val_reg, val_pc=None, lbl_fill=None, reg_f=rf, pc_f=pf):
        _cell(ws, ri, 1, label,   lbl_fill or yf if label else None)
        _cell(ws, ri, 2, val_reg, reg_f if label else None)
        if has_pc:
            _cell(ws, ri, 3, val_pc if val_pc is not None else '', pc_f if label else None)

    def _3d_block(arr_bil, arr_bic, prefix=''):
        """Costruisce le righe di una sezione 3D per una griglia."""
        v_bil = arr_bil[~np.isnan(arr_bil)] if arr_bil.size else np.array([])
        v_bic = arr_bic[~np.isnan(arr_bic)] if arr_bic.size else np.array([])
        rows = []
        if v_bil.size > 0:
            rows += [
                (f'{prefix}3D bilinear mean   (mm)', _r(np.mean(v_bil))),
                (f'{prefix}3D bilinear median (mm)', _r(np.median(v_bil))),
                (f'{prefix}3D bilinear max    (mm)', _r(np.max(v_bil))),
                (f'{prefix}3D bilinear min    (mm)', _r(np.min(v_bil))),
                (f'{prefix}3D bilinear N valid',     int(v_bil.size)),
                ('', ''),
            ]
        else:
            rows += [(f'{prefix}3D bilinear (mm)', 'N/A'), ('', '')]
        if v_bic.size > 0:
            rows += [
                (f'{prefix}3D bicubic  mean   (mm)', _r(np.mean(v_bic))),
                (f'{prefix}3D bicubic  median (mm)', _r(np.median(v_bic))),
                (f'{prefix}3D bicubic  max    (mm)', _r(np.max(v_bic))),
                (f'{prefix}3D bicubic  min    (mm)', _r(np.min(v_bic))),
                (f'{prefix}3D bicubic  N valid',     int(v_bic.size)),
                ('', ''),
            ]
        else:
            rows += [(f'{prefix}3D bicubic (mm)', 'N/A'), ('', '')]
        both = (~np.isnan(arr_bil)) & (~np.isnan(arr_bic))
        if both.any():
            dm = float(np.mean(arr_bic[both] - arr_bil[both]))
            dmed = float(np.median(arr_bic[both] - arr_bil[both]))
            rows += [
                (f'{prefix}Δ (bicubic-bilinear) mean   (mm)', _r(dm)),
                (f'{prefix}Δ (bicubic-bilinear) median (mm)', _r(dmed)),
                (f'{prefix}Best method (lower mean)',
                 'bicubic' if dm < 0 else ('bilinear' if dm > 0 else 'tie')),
            ]
        return rows

    ri = 2

    # ── Sezione comune (2D) ─────────────────────────────────────────────────
    common_rows = [
        ('N° ArUco corners evaluated',  n_corners),
        ('Scale factor px/mm (HSI)',     round(px_per_mm, 4)),
        ('', ''),
        ('2D mean   error (px)',         _r(_stat(err2d_px, np.mean))),
        ('2D median error (px)',         _r(_stat(err2d_px, np.median))),
        ('2D max    error (px)',         _r(_stat(err2d_px, np.max))),
        ('2D min    error (px)',         _r(_stat(err2d_px, np.min))),
        ('', ''),
        ('2D mean   error (mm)',         _r(_stat(err2d_mm_arr, np.mean))),
        ('2D median error (mm)',         _r(_stat(err2d_mm_arr, np.median))),
        ('2D max    error (mm)',         _r(_stat(err2d_mm_arr, np.max))),
        ('2D min    error (mm)',         _r(_stat(err2d_mm_arr, np.min))),
        ('', ''),
    ]
    for label, val in common_rows:
        _sum_row(ws2, ri, label, val, val)   # stesso valore in entrambe le colonne
        ri += 1

    # ── Sezione 3D REG ──────────────────────────────────────────────────────
    _hdr(ws2, ri, 1, f'3D errors on{reg_lbl}',
         fill=PatternFill('solid', start_color='2E75B6'))
    if has_pc:
        ws2.cell(row=ri, column=3, value='')
    ri += 1

    reg_3d_rows = _3d_block(err3d_bil, err3d_bic)
    for label, val in reg_3d_rows:
        _sum_row(ws2, ri, label, val, None, lbl_fill=yf, pc_f=None)
        ri += 1

    # ── Sezione 3D PC ───────────────────────────────────────────────────────
    if has_pc:
        _hdr(ws2, ri, 1, f'3D errors on{pc_lbl}',
             fill=PatternFill('solid', start_color='375623'))
        ws2.cell(row=ri, column=2, value='')
        ws2.cell(row=ri, column=3, value='')
        ri += 1

        pc_3d_rows = _3d_block(err3d_pc_bil, err3d_pc_bic)
        for label, val in pc_3d_rows:
            _sum_row(ws2, ri, label, None, val, lbl_fill=yf, reg_f=None)
            ri += 1

        # ── Sezione Δ (PC - REG) ─────────────────────────────────────────
        _hdr(ws2, ri, 1, f'Δ PC − REG  (negativo = PC migliore)',
             fill=PatternFill('solid', start_color='833C00'))
        ws2.cell(row=ri, column=2, value='')
        ws2.cell(row=ri, column=3, value='')
        ri += 1

        def _delta_val(arr_pc, arr_reg):
            v_pc  = arr_pc [~np.isnan(arr_pc)]
            v_reg = arr_reg[~np.isnan(arr_reg)]
            n = min(len(v_pc), len(v_reg))
            if n == 0:
                return 'N/A'
            both = (~np.isnan(arr_pc)) & (~np.isnan(arr_reg))
            if not both.any():
                return 'N/A'
            return _r(float(np.mean(arr_pc[both] - arr_reg[both])))

        delta_rows = [
            ('Δ bilinear mean   (mm)',
             _delta_val(err3d_pc_bil, err3d_bil)),
            ('Δ bicubic  mean   (mm)',
             _delta_val(err3d_pc_bic, err3d_bic)),
        ]
        for label, val in delta_rows:
            c = ws2.cell(row=ri, column=1, value=label)
            c.font, c.fill, c.alignment, c.border = norm_font, df, cen, bd
            c2 = ws2.cell(row=ri, column=2, value='')
            c2.border = bd
            c3 = ws2.cell(row=ri, column=3, value=val)
            c3.font, c3.fill, c3.alignment, c3.border = bold_font, df, cen, bd
            ri += 1

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 22
    if has_pc:
        ws2.column_dimensions['C'].width = 22

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"[Excel] Saved: {output_path}")


# =============================================================================
# 9 — Spectral point cloud builder
# =============================================================================

def _export_csv_streaming(cube, xyz_flat, valid_idx, H_inv32, W_r,
                          cols_hsi, rows_hsi, wavelengths, output_path,
                          max_points, reflectance_norm):
    """Export CSV campionando casualmente max_points punti, in modo streaming."""
    n_valid = len(valid_idx)
    if n_valid > max_points:
        sel = np.sort(np.random.choice(n_valid, max_points, replace=False))
        valid_idx = valid_idx[sel]
    n_out  = len(valid_idx)
    bands  = cube.shape[2]
    col_r  = (valid_idx % W_r).astype(np.float32)
    row_r  = (valid_idx // W_r).astype(np.float32)
    pts    = np.column_stack([col_r, row_r])
    hsi    = cv2.perspectiveTransform(pts.reshape(-1, 1, 2), H_inv32).reshape(-1, 2)
    col_h  = np.clip(np.round(hsi[:, 0]).astype(np.int32), 0, cols_hsi - 1)
    row_h  = np.clip(np.round(hsi[:, 1]).astype(np.int32), 0, rows_hsi - 1)

    band_hdr = (
        [f"b{i}_{int(round(w))}nm" for i, w in enumerate(wavelengths)]
        if wavelengths and len(wavelengths) == bands
        else [f"b{i:03d}" for i in range(bands)]
    )
    header = "X_mm,Y_mm,Z_mm," + ",".join(band_hdr)
    chunk  = 50_000
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'w') as f:
        f.write(header + "\n")
        for s in range(0, n_out, chunk):
            e      = min(s + chunk, n_out)
            xyz_c  = xyz_flat[valid_idx[s:e]].astype(np.float32)
            sp_c   = cube[row_h[s:e], col_h[s:e], :].astype(np.float32)
            if reflectance_norm:
                sp_min = sp_c.min(axis=1, keepdims=True)
                sp_max = sp_c.max(axis=1, keepdims=True)
                sp_c   = (sp_c - sp_min) / np.where(sp_max > sp_min, sp_max - sp_min, 1.0)
            data = np.concatenate([xyz_c, sp_c], axis=1)
            np.savetxt(f, data, delimiter=',', fmt='%.4f')
    print(f"[CSV] Salvato: {output_path}  ({n_out:,} punti, "
          f"{os.path.getsize(output_path)/1e6:.1f} MB)")


def build_and_export_pointcloud_streaming(
    cube, xyz_map, H_inv,
    output_ply_path=None,
    output_npz_prefix=None,
    wavelengths=None,
    border_px=2,
    reflectance_norm=True,
    chunk_size=100_000,
    proj_chunk=500_000,
):
    """
    Costruisce ed esporta il point cloud spettrale in modalità streaming,
    mantenendo in RAM al massimo `chunk_size` spettri alla volta.

    Rispetto a build_spectral_pointcloud() + export_*():
      - NON accumula l'array completo (N, bands) in RAM
      - Peak RAM ≈ chunk_size × (3 + bands) × 4 byte  (es. 100K × 293 × 4 = ~117 MB)
      - Scrive il PLY direttamente chunk per chunk
      - Per NPZ salva file memmap (.bin) + metadati .npy ricaricabili senza caricare tutto
        Caricamento in Python:
            xyz     = np.memmap('<prefix>_xyz.bin',     dtype='float32',
                                mode='r', shape=(N, 3))
            spectra = np.memmap('<prefix>_spectra.bin', dtype='float32',
                                mode='r', shape=(N, bands))
            wl      = np.load('<prefix>_wavelengths.npy')

    Parameters
    ----------
    cube             : (rows_hsi, cols_hsi, bands) float32
    xyz_map          : (H_r, W_r, 3) float32  — coordinate 3D per pixel render
    H_inv            : (3,3) omografia inversa render_px -> HSI_px
    output_ply_path  : percorso PLY di output (None = skip)
    output_npz_prefix: prefisso per i file memmap (None = skip)
    wavelengths      : lista lunghezze d'onda in nm (opzionale)
    border_px        : margine di guardia sui bordi HSI
    reflectance_norm : normalizza ogni spettro in [0, 1]
    chunk_size       : righe di spettri per chunk (controlla il picco RAM)
    proj_chunk       : chunk per la fase di back-projection (può essere più grande)

    Returns
    -------
    valid_mask : (H_r, W_r) bool  — maschera copertura sul render
    n_valid    : int              — numero totale di punti nel cloud
    """
    H_r, W_r, _           = xyz_map.shape
    rows_hsi, cols_hsi, bands = cube.shape
    H_inv32 = H_inv.astype(np.float32)
    xyz_flat = xyz_map.reshape(-1, 3)   # view, nessuna copia
    N_total  = H_r * W_r

    print(f"\n[BuildPC streaming] Render grid : {W_r}×{H_r}  ({N_total:,} px)")
    print(f"[BuildPC streaming] HSI image   : {cols_hsi}×{rows_hsi}  x {bands} bands")
    print(f"[BuildPC streaming] Chunk size  : {chunk_size:,} punti")

    # ── Pass 1: back-project tutti i pixel mesh-validi -> trova quelli in FOV ──
    # Accumula in liste per chunk, poi concatena.
    # Peak RAM: due liste di array piccoli, al massimo proj_chunk elementi ciascuna.
    print("[BuildPC streaming] Pass 1: rilevamento punti validi...")

    valid_global_idx_list = []   # indice flat nel render
    hsi_col_list          = []   # colonna HSI (int32)
    hsi_row_list          = []   # riga HSI (int32)
    n_mesh_hit            = 0
    n_in_fov              = 0

    for s in range(0, N_total, proj_chunk):
        e         = min(s + proj_chunk, N_total)
        mesh_local = ~np.isnan(xyz_flat[s:e, 0])  # bool (chunk,)
        if not mesh_local.any():
            continue
        n_mesh_hit += int(mesh_local.sum())

        local_hit  = np.where(mesh_local)[0]
        global_hit = (s + local_hit).astype(np.int32)

        col_r = (global_hit % W_r).astype(np.float32)
        row_r = (global_hit // W_r).astype(np.float32)
        pts   = np.column_stack([col_r, row_r])
        hsi   = cv2.perspectiveTransform(
            pts.reshape(-1, 1, 2), H_inv32
        ).reshape(-1, 2)

        col_h = hsi[:, 0]
        row_h = hsi[:, 1]
        in_fov = ((col_h >= border_px) & (col_h <= cols_hsi - 1 - border_px) &
                  (row_h >= border_px) & (row_h <= rows_hsi - 1 - border_px))
        if not in_fov.any():
            continue
        n_in_fov += int(in_fov.sum())

        valid_global_idx_list.append(global_hit[in_fov])
        hsi_col_list.append(np.round(col_h[in_fov]).astype(np.int32))
        hsi_row_list.append(np.round(row_h[in_fov]).astype(np.int32))

        print(f"  scansionato {e:,}/{N_total:,}  mesh_hit={n_mesh_hit:,}  "
              f"in_fov={n_in_fov:,}", end='\r')

    print()
    print(f"[BuildPC streaming] Pixels con hit mesh: {n_mesh_hit:,}")
    print(f"[BuildPC streaming] Pixels in HSI FOV  : {n_in_fov:,}")

    if n_in_fov == 0:
        print("[BuildPC streaming] Nessun punto valido trovato.")
        return np.zeros((H_r, W_r), dtype=bool), 0

    # Concatena e libera le liste
    valid_idx = np.concatenate(valid_global_idx_list)  # (n_valid,) int32
    hsi_col   = np.concatenate(hsi_col_list)           # (n_valid,) int32
    hsi_row   = np.concatenate(hsi_row_list)           # (n_valid,) int32
    del valid_global_idx_list, hsi_col_list, hsi_row_list

    n_valid    = len(valid_idx)
    valid_flat = np.zeros(N_total, dtype=bool)
    valid_flat[valid_idx] = True
    valid_mask = valid_flat.reshape(H_r, W_r)

    # Clamp indici HSI per sicurezza
    hsi_col = np.clip(hsi_col, 0, cols_hsi - 1)
    hsi_row = np.clip(hsi_row, 0, rows_hsi - 1)

    n_chunks = (n_valid + chunk_size - 1) // chunk_size

    # ── PLY streaming ─────────────────────────────────────────────────────────
    if output_ply_path is not None:
        os.makedirs(os.path.dirname(os.path.abspath(output_ply_path)), exist_ok=True)
        band_names = (
            [f"b{i}_{int(round(w))}nm" for i, w in enumerate(wavelengths)]
            if wavelengths and len(wavelengths) == bands
            else [f"b{i:03d}" for i in range(bands)]
        )
        header = (
            "ply\nformat binary_little_endian 1.0\n"
            f"element vertex {n_valid}\n"
            "property float x\nproperty float y\nproperty float z\n"
            + "\n".join(f"property float {n}" for n in band_names)
            + "\nend_header\n"
        )
        print(f"[BuildPC streaming] Scrittura PLY in {n_chunks} chunk "
              f"-> {output_ply_path}")
        with open(output_ply_path, 'wb') as f_ply:
            f_ply.write(header.encode('ascii'))
            for ci, s in enumerate(range(0, n_valid, chunk_size)):
                e  = min(s + chunk_size, n_valid)
                xyz_c     = xyz_flat[valid_idx[s:e]].astype(np.float32)
                sp_c      = cube[hsi_row[s:e], hsi_col[s:e], :].astype(np.float32)
                if reflectance_norm:
                    sp_min = sp_c.min(axis=1, keepdims=True)
                    sp_max = sp_c.max(axis=1, keepdims=True)
                    sp_c   = (sp_c - sp_min) / np.where(sp_max > sp_min, sp_max - sp_min, 1.0)
                f_ply.write(np.concatenate([xyz_c, sp_c], axis=1).tobytes())
                print(f"  PLY chunk {ci+1}/{n_chunks}  [{e:,}/{n_valid:,}]", end='\r')
        print(f"\n[BuildPC streaming] PLY salvato  "
              f"({os.path.getsize(output_ply_path)/1e9:.2f} GB)")

    # ── NPZ via memmap (zero picco RAM) ───────────────────────────────────────
    if output_npz_prefix is not None:
        os.makedirs(os.path.dirname(os.path.abspath(output_npz_prefix + '_xyz.bin')),
                    exist_ok=True)
        xyz_path     = output_npz_prefix + '_xyz.bin'
        spectra_path = output_npz_prefix + '_spectra.bin'
        wl_path      = output_npz_prefix + '_wavelengths.npy'
        meta_path    = output_npz_prefix + '_meta.npy'

        print(f"[BuildPC streaming] Scrittura memmap NPZ in {n_chunks} chunk...")
        xyz_mm  = np.memmap(xyz_path,     dtype=np.float32, mode='w+',
                            shape=(n_valid, 3))
        spec_mm = np.memmap(spectra_path, dtype=np.float32, mode='w+',
                            shape=(n_valid, bands))

        for ci, s in enumerate(range(0, n_valid, chunk_size)):
            e        = min(s + chunk_size, n_valid)
            xyz_mm[s:e]  = xyz_flat[valid_idx[s:e]].astype(np.float32)
            sp_c         = cube[hsi_row[s:e], hsi_col[s:e], :].astype(np.float32)
            if reflectance_norm:
                sp_min  = sp_c.min(axis=1, keepdims=True)
                sp_max  = sp_c.max(axis=1, keepdims=True)
                sp_c    = (sp_c - sp_min) / np.where(sp_max > sp_min, sp_max - sp_min, 1.0)
            spec_mm[s:e] = sp_c
            print(f"  NPZ chunk {ci+1}/{n_chunks}  [{e:,}/{n_valid:,}]", end='\r')

        del xyz_mm, spec_mm   # flush su disco
        np.save(wl_path,   np.array(wavelengths or [], dtype=np.float32))
        np.save(meta_path, np.array([n_valid, 3, n_valid, bands], dtype=np.int64))
        print(f"\n[BuildPC streaming] Memmap salvati in {output_npz_prefix}*.bin")
        print(f"  Caricamento:  np.memmap('{output_npz_prefix}_xyz.bin', "
              f"dtype='float32', mode='r', shape=({n_valid}, 3))")

    print(f"[BuildPC streaming] Completato: {n_valid:,} punti  x {bands} bande")
    return valid_mask, n_valid


def parse_wavelengths(meta):
    """Extract wavelength list (nm) from ENVI header metadata. Returns None if absent."""
    wl_str = meta.get('wavelength', None)
    if wl_str is None:
        return None
    try:
        return [float(w) for w in wl_str.strip('{}').strip().split(',')]
    except ValueError:
        return None


def build_spectral_pointcloud(cube, xyz_map, H_inv,
                               border_px=2, reflectance_norm=True):
    """
    For every render pixel with a valid (X,Y,Z), back-project via H_inv to
    the HSI image and read the full spectrum.

    Parameters
    ----------
    cube             : (rows_hsi, cols_hsi, bands) float32
    xyz_map          : (H_r, W_r, 3) float32  3D coords per render pixel (mm)
    H_inv            : (3,3) inverse homography: render pixel -> HSI pixel
    border_px        : discard pixels within this many pixels of the HSI edge
    reflectance_norm : if True, normalise each spectrum to [0, 1]

    Returns
    -------
    points_xyz : (N, 3)       float32  X, Y, Z in mm
    spectra    : (N, bands)   float32
    valid_mask : (H_r, W_r)  bool     coverage mask on render
    n_valid    : int
    """
    H_r, W_r, _           = xyz_map.shape
    rows_hsi, cols_hsi, bands = cube.shape

    print(f"\n[BuildPC] Render grid : {W_r}x{H_r} px  ({W_r*H_r:,} pixels)")
    print(f"[BuildPC] HSI image   : {cols_hsi}x{rows_hsi} px  x {bands} bands")

    col_grid, row_grid = np.meshgrid(np.arange(W_r), np.arange(H_r))
    render_coords = np.column_stack([
        col_grid.ravel().astype(np.float32),
        row_grid.ravel().astype(np.float32),
    ])

    xyz_flat   = xyz_map.reshape(-1, 3)
    mesh_valid = ~np.isnan(xyz_flat[:, 0])
    print(f"[BuildPC] Pixels with mesh hit: {mesh_valid.sum():,} / {mesh_valid.size:,}")

    pts_hsi_back = cv2.perspectiveTransform(
        render_coords[mesh_valid].reshape(-1, 1, 2).astype(np.float32),
        H_inv.astype(np.float32)
    ).reshape(-1, 2)

    col_hsi = pts_hsi_back[:, 0]
    row_hsi = pts_hsi_back[:, 1]

    in_bounds = (
        (col_hsi >= border_px) & (col_hsi <= cols_hsi - 1 - border_px) &
        (row_hsi >= border_px) & (row_hsi <= rows_hsi - 1 - border_px)
    )
    print(f"[BuildPC] Pixels inside HSI FOV: {in_bounds.sum():,}")

    col_idx = np.round(col_hsi[in_bounds]).astype(int)
    row_idx = np.round(row_hsi[in_bounds]).astype(int)

    xyz_in_bounds = xyz_flat[mesh_valid][in_bounds]

    print("[BuildPC] Reading spectra from cube...")
    spectra = cube[row_idx, col_idx, :].astype(np.float32)

    if reflectance_norm:
        s_min = spectra.min(axis=1, keepdims=True)
        s_max = spectra.max(axis=1, keepdims=True)
        spectra = (spectra - s_min) / np.where(s_max > s_min, s_max - s_min, 1.0)
        print("[BuildPC] Spectra normalised to [0, 1] per point.")

    mesh_valid_idx   = np.where(mesh_valid)[0]
    in_bounds_global = mesh_valid_idx[in_bounds]
    valid_flat       = np.zeros(H_r * W_r, dtype=bool)
    valid_flat[in_bounds_global] = True
    valid_mask = valid_flat.reshape(H_r, W_r)

    n_valid = xyz_in_bounds.shape[0]
    print(f"[BuildPC] Final cloud: {n_valid:,} points  x {bands} bands")
    return xyz_in_bounds.astype(np.float32), spectra, valid_mask, n_valid


# =============================================================================
# 10 — Export point cloud
# =============================================================================

def export_ply(points_xyz, spectra, output_path, wavelengths=None):
    """
    Export spectral point cloud as a binary PLY file.
    Open in CloudCompare and colour by any scalar field (band).
    """
    N, bands = spectra.shape
    band_names = ([f"b{i}_{int(round(w))}nm" for i, w in enumerate(wavelengths)]
                  if wavelengths and len(wavelengths) == bands
                  else [f"b{i:03d}" for i in range(bands)])
    header = ("ply\nformat binary_little_endian 1.0\n"
              f"element vertex {N}\n"
              "property float x\nproperty float y\nproperty float z\n"
              + "\n".join(f"property float {n}" for n in band_names)
              + "\nend_header\n")
    data = np.concatenate([points_xyz, spectra], axis=1)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, 'wb') as f:
        f.write(header.encode('ascii'))
        f.write(data.astype(np.float32).tobytes())
    print(f"[Export PLY] {output_path}  ({N:,} pts, {os.path.getsize(output_path)/1e6:.1f} MB)")


def export_npy(points_xyz, spectra, output_path, wavelengths=None):
    """
    Export as compressed numpy .npz.

    Load with:
        data = np.load('spectral_pointcloud.npz')
        xyz     = data['xyz']         # (N, 3)  mm
        spectra = data['spectra']     # (N, bands)
        wl      = data['wavelengths'] # (bands,) nm
    """
    save_dict = {'xyz': points_xyz, 'spectra': spectra}
    if wavelengths is not None:
        save_dict['wavelengths'] = np.array(wavelengths, dtype=np.float32)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    np.savez_compressed(output_path, **save_dict)
    print(f"[Export NPZ] {output_path}.npz  ({os.path.getsize(output_path+'.npz')/1e6:.1f} MB)")


def export_csv(points_xyz, spectra, output_path, wavelengths=None,
               max_points=500_000):
    """
    Export as CSV, capped at max_points (sub-sampled randomly if exceeded).
    Columns: X_mm, Y_mm, Z_mm, b0, b1, ..., bN
    """
    N, bands = spectra.shape
    if N > max_points:
        idx = np.random.choice(N, max_points, replace=False)
        points_xyz, spectra = points_xyz[idx], spectra[idx]
    band_hdr = ([f"b{i}_{int(round(w))}nm" for i, w in enumerate(wavelengths)]
                if wavelengths and len(wavelengths) == bands
                else [f"b{i:03d}" for i in range(bands)])
    header = "X_mm,Y_mm,Z_mm," + ",".join(band_hdr)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    np.savetxt(output_path, np.concatenate([points_xyz, spectra], axis=1),
               delimiter=',', header=header, comments='', fmt='%.4f')
    print(f"[Export CSV] {output_path}  ({points_xyz.shape[0]:,} pts, "
          f"{os.path.getsize(output_path)/1e6:.1f} MB)")


# =============================================================================
# 11 — Visualisations
# =============================================================================

def visualize_registration(render_rgb, hsi_2d, pts_render_h, data_hsi, data_render,
                            suspicious_pixels_hsi=None, coords_3d_h=None,
                            output_path='output_registration_3D.png'):
    """
    Side-by-side image: HSI (left) and render (right) with ArUco overlays
    and projected suspicious points.
    """
    panel = render_rgb.copy().astype(np.uint8)
    if pts_render_h is not None:
        ov = panel.copy()
        for px, py in pts_render_h:
            cv2.circle(ov, (int(round(px)), int(round(py))), 8, (0,0,255), -1)
        panel = cv2.addWeighted(panel, 0.5, ov, 0.5, 0)
        for i, (px, py) in enumerate(pts_render_h):
            lbl = f"T{i}"
            if coords_3d_h is not None and not np.isnan(coords_3d_h[i,0]):
                lbl += f" Z={coords_3d_h[i,2]:.1f}"
            cv2.putText(panel, lbl, (int(round(px))+8, int(round(py))-8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0,0,255), 1)
    if data_render:
        for mid, corners in data_render.items():
            ci = corners.astype(int)
            for j in range(4):
                cv2.line(panel, tuple(ci[j]), tuple(ci[(j+1)%4]), (0,220,255), 1)
            cx, cy = int(ci[:,0].mean()), int(ci[:,1].mean())
            cv2.putText(panel, f"ID{mid}", (cx+4,cy-4),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,255,0), 1)
    cv2.putText(panel, "RENDER top-view", (10,20), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 1)

    if hsi_2d is not None:
        hsi_panel = cv2.cvtColor(hsi_2d, cv2.COLOR_GRAY2BGR)
        sh = panel.shape[0] / hsi_panel.shape[0]
        if suspicious_pixels_hsi is not None:
            ov_h = hsi_panel.copy()
            for px, py in suspicious_pixels_hsi:
                cv2.circle(ov_h, (int(round(px)),int(round(py))),
                           max(1,int(round(8/sh))), (0,0,255), -1)
            hsi_panel = cv2.addWeighted(hsi_panel, 0.5, ov_h, 0.5, 0)
        if data_hsi:
            for mid, corners in data_hsi.items():
                ci = corners.astype(int)
                for j in range(4):
                    cv2.line(hsi_panel, tuple(ci[j]), tuple(ci[(j+1)%4]), (0,220,255), 1)
                cx, cy = int(ci[:,0].mean()), int(ci[:,1].mean())
                cv2.putText(hsi_panel, f"ID{mid}", (cx+4,cy-4),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,0), 1)
        cv2.putText(hsi_panel, "HSI", (10,20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255,255,255), 2)
        hsi_panel = cv2.resize(hsi_panel, (int(hsi_panel.shape[1]*sh), panel.shape[0]),
                               interpolation=cv2.INTER_LINEAR)
        panel = np.hstack([hsi_panel, panel])

    cv2.imwrite(output_path, panel)
    print(f"[Viz] Saved: {output_path}")
    return panel


def visualize_coverage(render_rgb, valid_mask, hsi_2d, H_inv,
                        output_path='coverage_map.png'):
    """
    Side-by-side image:
      Left  — HSI with the region covered by the point cloud (green overlay)
      Right — Render with valid 3D+spectral pixels highlighted (green overlay)
    """
    panel = render_rgb.copy()
    ov = panel.copy()
    ov[valid_mask] = [0,200,0]
    panel = cv2.addWeighted(panel, 0.4, ov, 0.6, 0)
    cv2.putText(panel, "RENDER — valid 3D+spectral pixels",
                (10,20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,255,255), 1)

    hsi_panel = cv2.cvtColor(hsi_2d, cv2.COLOR_GRAY2BGR)
    rows_hsi, cols_hsi = hsi_2d.shape
    vr, vc = np.where(valid_mask)
    pts_r  = np.column_stack([vc, vr]).astype(np.float32)
    if len(pts_r) > 0:
        step    = max(1, len(pts_r) // 50_000)
        pts_h   = cv2.perspectiveTransform(pts_r[::step].reshape(-1,1,2),
                                           H_inv.astype(np.float32)).reshape(-1,2).astype(int)
        ok = ((pts_h[:,0]>=0)&(pts_h[:,0]<cols_hsi)&
              (pts_h[:,1]>=0)&(pts_h[:,1]<rows_hsi))
        for cx, cy in pts_h[ok]:
            cv2.circle(hsi_panel, (cx,cy), 1, (0,200,0), -1)
    cv2.putText(hsi_panel, "HSI — region mapped to 3D",
                (10,20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,255,255), 1)

    sh   = panel.shape[0] / hsi_panel.shape[0]
    hsi_panel = cv2.resize(hsi_panel,
                           (int(hsi_panel.shape[1]*sh), panel.shape[0]),
                           interpolation=cv2.INTER_LINEAR)
    combined = np.hstack([hsi_panel, panel])
    cv2.imwrite(output_path, combined)
    print(f"[Coverage] Saved: {output_path}")
    return combined


def save_turbo_render(xyz_map, output_path):
    """Save a TURBO-colourmap depth image from the xyz_map Z channel."""
    z_map   = xyz_map[:,:,2]
    valid_z = ~np.isnan(z_map)
    z_norm  = np.zeros(z_map.shape, dtype=np.float32)
    if valid_z.any():
        z_min, z_max = z_map[valid_z].min(), z_map[valid_z].max()
        z_norm[valid_z] = (z_map[valid_z]-z_min) / max(z_max-z_min, 1e-8)
    render_color = cv2.applyColorMap((z_norm*255).astype(np.uint8), cv2.COLORMAP_TURBO)
    render_color[~valid_z] = 0
    cv2.imwrite(output_path, render_color)
    print(f"[Render TURBO] Saved: {output_path}")


# =============================================================================
# 12 — Clustering helper
# =============================================================================

def extract_suspicious_centroids(suspicious_mask, min_area=5):
    """
    Find connected components in a binary mask and return their centroids.

    Parameters
    ----------
    suspicious_mask : (H, W) uint8 or bool  binary mask
    min_area        : minimum component area in pixels to keep

    Returns
    -------
    centroids : (N, 2) float32  [col, row] per cluster
    n         : int             number of clusters found
    """
    mask_u8 = suspicious_mask.astype(np.uint8)
    n, _, stats, cents = cv2.connectedComponentsWithStats(mask_u8, connectivity=8)
    result = []
    for lid in range(1, n):
        area = stats[lid, cv2.CC_STAT_AREA]
        if area < min_area:
            continue
        result.append([cents[lid,0], cents[lid,1]])
        print(f"  [Clustering] {lid}: area={area}px  "
              f"centroid=({cents[lid,0]:.1f},{cents[lid,1]:.1f})")
    if not result:
        return np.empty((0,2), dtype=np.float32), 0
    arr = np.array(result, dtype=np.float32)
    print(f"[Clustering] {len(arr)} region(s) found.")
    return arr, len(arr)


# =============================================================================
# MAIN PIPELINE
# =============================================================================

def run_full_pipeline(
    hsi_hdr_path,
    mesh_path,
    aruco_json_path,
    output_dir,
    # Registration
    hsi_extraction_method  = 'mean',
    aruco_dict_type        = cv2.aruco.DICT_4X4_50,
    suspicious_pixels_hsi  = None,
    render_resolution_mm   = 2.0,      # risoluzione point cloud
    render_resolution_reg_mm = None,   # risoluzione registrazione (None = uguale a render_resolution_mm)
    render_margin_mm       = 10.0,
    chunk_size             = 500_000,
    marker_side_mm         = 9.0,
    use_subpix             = True,     # subpixel refinement ArUco (Punto 3)
    subpix_winsize         = 5,
    # Point cloud
    border_px              = 2,
    reflectance_norm       = True,
    export_ply_file        = True,
    export_npy_file        = True,
    export_csv_file        = False,
    csv_max_points         = 500_000,
    pc_chunk_size          = 100_000,  # righe spettri per chunk (streaming)
    # Export flags
    save_pointcloud        = True,     # False = salta costruzione e salvataggio point cloud
    save_images            = True,     # False = non salva nessuna immagine (render, overlay, coverage)
    sample_name            = 'sample',
    precomputed_render     = None,     # render point cloud pre-calcolato (GPU)
                                       # tuple (render_rgb, depth_map, xyz_map, origin_xy, res)
    precomputed_render_reg = None,     # render registrazione pre-calcolato (GPU, res più grossolana)
                                       # stessa struttura; se None e precomputed_render fornito,
                                       # usa precomputed_render anche per la registrazione
):
    """
    Full pipeline:
      1. 3D registration  (HSI -> mesh via ArUco + homography)
      2. Excel error report  (2D px, 2D mm, 3D bil/bic mm)
      3. Spectral point cloud  (PLY + NPZ + optional CSV)
      4. Visualisations  (registration overlay, coverage map, TURBO depth)

    Doppio render (Punto 1):
      - render_resolution_reg_mm  : risoluzione usata SOLO per l'omografia
                                    (default = render_resolution_mm).
                                    Un valore più grossolano (es. 0.5 mm/px)
                                    garantisce 12/12 inlier RANSAC con la
                                    detection ArUco.
      - render_resolution_mm      : risoluzione usata per il point cloud
                                    e il lookup 3D (es. 0.3 mm/px per alta densità).
      Se le due risoluzioni coincidono viene fatto un solo render.

    SubPix (Punto 3):
      - use_subpix / subpix_winsize : abilita cv2.cornerSubPix dopo la
                                      detection ArUco sull'HSI.

    Parameters
    ----------
    hsi_hdr_path             : ENVI .hdr file path
    mesh_path                : 3D mesh (.ply / .obj)
    aruco_json_path          : JSON with ArUco 3D corners in metres
    output_dir               : folder where all outputs are saved
    hsi_extraction_method    : 'mean' | 'visible_band' | 'pca'
    aruco_dict_type          : ArUco dictionary constant
    suspicious_pixels_hsi    : (N, 2) float32  HSI pixel coordinates to locate in 3D
    render_resolution_mm     : mm/px point cloud render
    render_resolution_reg_mm : mm/px registration render (None = same as above)
    render_margin_mm         : padding around mesh bounding box in mm
    chunk_size               : rays per batch for CPU render
    marker_side_mm           : physical side length of ArUco markers in mm
    use_subpix               : enable sub-pixel corner refinement
    subpix_winsize           : half-window size for cornerSubPix
    border_px                : HSI edge guard in pixels
    reflectance_norm         : normalise each spectrum to [0, 1]
    export_ply_file          : save .ply
    export_npy_file          : save .npz
    export_csv_file          : save .csv
    csv_max_points           : max points in CSV
    sample_name              : prefix for all output filenames
    precomputed_render       : render point cloud GPU pre-calcolato
    precomputed_render_reg   : render registrazione GPU pre-calcolato

    Returns
    -------
    dict with keys:
      'H', 'H_inv', 'H_pc', 'H_pc_inv',
      'pts_hsi', 'pts_render', 'common_ids',
      'data_hsi', 'data_render', 'err2d_px', 'err3d_mm', 'err3d_dict',
      'best_method', 'pts_render_h', 'coords_3d_h',
      'coords_3d_h_bilinear', 'coords_3d_h_bicubic',
      'render_rgb', 'depth_map', 'xyz_map', 'origin_xy',
      'render_rgb_reg', 'depth_map_reg', 'xyz_map_reg', 'origin_xy_reg',
      'cube', 'hsi_2d', 'wavelengths',
      'xyz', 'spectra', 'valid_mask'
    """
    os.makedirs(output_dir, exist_ok=True)

    # Normalizza la risoluzione di registrazione
    res_reg = render_resolution_reg_mm if render_resolution_reg_mm is not None \
              else render_resolution_mm
    res_pc  = render_resolution_mm
    dual    = abs(res_reg - res_pc) > 1e-6   # True = due render distinti

    print("=" * 68)
    print("FULL PIPELINE: 3D REGISTRATION + SPECTRAL POINT CLOUD")
    print(f"  Sample         : {sample_name}")
    print(f"  Res. registro  : {res_reg} mm/px")
    print(f"  Res. point cloud: {res_pc} mm/px")
    print(f"  SubPix refine  : {'ON' if use_subpix else 'OFF'}")
    print(f"  Output dir     : {output_dir}")
    print("=" * 68)

    # ── Step 0: Load ArUco 3D ─────────────────────────────────────────────────
    print("\n[Step 0] Loading ArUco 3D from JSON...")
    aruco_3d, _ = load_aruco_3d_json(aruco_json_path, scale_m_to_mm=True)

    # ── Step 1: Load HSI ──────────────────────────────────────────────────────
    print("\n[Step 1] Loading HSI...")
    cube, meta = load_envi(hsi_hdr_path)
    hsi_2d     = extract_2d_from_hsi(cube, meta, method=hsi_extraction_method)
    print(f"  Cube shape: {cube.shape}")

    # ── Step 2: Render(s) ─────────────────────────────────────────────────────
    # Render REGISTRAZIONE (risoluzione grossolana, massimizza inlier RANSAC)
    if precomputed_render_reg is not None:
        print("\n[Step 2a] Render registrazione pre-calcolato (GPU).")
        render_rgb_reg, depth_map_reg, xyz_map_reg, origin_xy_reg, res_reg_actual = \
            precomputed_render_reg
    elif precomputed_render is not None and not dual:
        # Un solo render pre-calcolato, usalo per entrambi
        print("\n[Step 2a] Render unico pre-calcolato — usato per registrazione.")
        render_rgb_reg, depth_map_reg, xyz_map_reg, origin_xy_reg, res_reg_actual = \
            precomputed_render
    else:
        print(f"\n[Step 2a] Render registrazione CPU ({res_reg} mm/px)...")
        mesh = load_mesh(mesh_path, scale_m_to_mm=True)
        render_rgb_reg, depth_map_reg, xyz_map_reg, origin_xy_reg, res_reg_actual = \
            render_orthographic_topview(
                mesh,
                resolution_mm_per_px = res_reg,
                margin_mm             = render_margin_mm,
                chunk_size            = chunk_size,
            )
        if save_images:
            save_render(render_rgb_reg, depth_map_reg, output_dir=output_dir,
                        prefix=f'{sample_name}_render_reg')
            save_turbo_render(xyz_map_reg,
                              os.path.join(output_dir,
                                           f'{sample_name}_render_reg_turbo.png'))

    # Render POINT CLOUD (risoluzione fine, massimizza densità)
    if not dual:
        # Stessa risoluzione → stesso render
        render_rgb_pc  = render_rgb_reg
        depth_map_pc   = depth_map_reg
        xyz_map_pc     = xyz_map_reg
        origin_xy_pc   = origin_xy_reg
        res_pc_actual  = res_reg_actual
        print(f"[Step 2b] Risoluzione registrazione = point cloud ({res_pc} mm/px) — render condiviso.")
    elif precomputed_render is not None:
        print("\n[Step 2b] Render point cloud pre-calcolato (GPU).")
        render_rgb_pc, depth_map_pc, xyz_map_pc, origin_xy_pc, res_pc_actual = \
            precomputed_render
    else:
        print(f"\n[Step 2b] Render point cloud CPU ({res_pc} mm/px)...")
        if 'mesh' not in dir():
            mesh = load_mesh(mesh_path, scale_m_to_mm=True)
        render_rgb_pc, depth_map_pc, xyz_map_pc, origin_xy_pc, res_pc_actual = \
            render_orthographic_topview(
                mesh,
                resolution_mm_per_px = res_pc,
                margin_mm             = render_margin_mm,
                chunk_size            = chunk_size,
            )
        if save_images:
            save_render(render_rgb_pc, depth_map_pc, output_dir=output_dir,
                        prefix=f'{sample_name}_render_pc')
            save_turbo_render(xyz_map_pc,
                              os.path.join(output_dir,
                                           f'{sample_name}_render_pc_turbo.png'))

    # ── Step 3: ArUco on HSI ──────────────────────────────────────────────────
    print("\n[Step 3] Detecting ArUco on HSI...")
    data_hsi, _ = detect_aruco(hsi_2d, aruco_dict_type=aruco_dict_type,
                                use_subpix=use_subpix,
                                subpix_winsize=subpix_winsize)
    print(f"  HSI markers found: {sorted(data_hsi.keys())}")

    # ── Step 3b: Project JSON corners -> render REGISTRAZIONE ─────────────────
    print("\n[Step 3b] Projecting JSON ArUco corners -> render pixels (reg)...")
    data_render_reg = project_aruco3d_to_render(
        aruco_3d, origin_xy_reg, res_reg_actual
    )

    # ── Step 4: Omografia su render REGISTRAZIONE ─────────────────────────────
    print("\n[Step 4] Computing homography HSI -> render_reg...")
    pts_hsi, pts_render_reg, common_ids = match_aruco_hsi_render(
        data_hsi, data_render_reg
    )
    H, _ = compute_homography(pts_hsi, pts_render_reg, tag=' HSI->render_reg')

    # H mappa HSI-px → render_reg-px.
    # Per il point cloud serve H_pc (HSI-px → render_pc-px).
    # Le due griglie hanno la stessa origin_xy (stessa mesh, stesso margin),
    # quindi il passaggio è una semplice scalatura:
    #   render_pc_col = render_reg_col * (res_reg / res_pc)
    # ovvero H_pc = S @ H  dove S = diag(s, s, 1), s = res_reg/res_pc
    scale_r2p = res_reg_actual / res_pc_actual          # > 1 se reg più grossolana
    S = np.array([[scale_r2p, 0,         0],
                  [0,         scale_r2p, 0],
                  [0,         0,         1]], dtype=np.float64)
    H_pc     = S @ H                                    # HSI-px → render_pc-px
    H_inv    = np.linalg.inv(H)                         # render_reg-px → HSI-px
    H_pc_inv = np.linalg.inv(H_pc)                     # render_pc-px  → HSI-px

    # ── Step 5: Suspicious points (su render POINT CLOUD) ────────────────────
    pts_render_h     = None
    coords_3d_h_bil  = None
    coords_3d_h_bic  = None
    if suspicious_pixels_hsi is not None and len(suspicious_pixels_hsi) > 0:
        print(f"\n[Step 5] Projecting {len(suspicious_pixels_hsi)} suspicious point(s)...")
        pts_render_h = cv2.perspectiveTransform(
            suspicious_pixels_hsi.reshape(-1, 1, 2).astype(np.float32), H_pc
        ).reshape(-1, 2)
        coords_3d_h_bil, coords_3d_h_bic = lookup_3d(
            pts_render_h, xyz_map_pc, mode='both', verbose=True
        )

    # ── Step 6: Error metrics ─────────────────────────────────────────────────
    # 2D sempre su render_reg (H è stata fittata su quella griglia).
    # 3D calcolato su ENTRAMBE le griglie:
    #   - reg: xyz_map_reg + H   → misura coerente con il fitting
    #   - pc : xyz_map_pc  + H_pc → misura sulla griglia del point cloud (0.3 mm/px)
    print("\n[Step 6] Computing registration errors...")
    err2d_px   = reprojection_error_2d(
        pts_hsi, pts_render_reg, H, tag='Homography'
    )
    err3d_dict = reprojection_error_3d_json(
        aruco_3d, data_hsi, data_render_reg, xyz_map_reg, H, tag='reg'
    )

    # Errori 3D sulla griglia PC (solo se le due griglie sono distinte)
    if dual:
        print("\n[Step 6b] Computing 3D errors on PC grid...")
        err3d_dict_pc = reprojection_error_3d_json(
            aruco_3d, data_hsi, data_render_reg, xyz_map_pc, H_pc, tag='pc'
        )
    else:
        # Griglia unica: gli errori sono identici, non ricalcolare
        err3d_dict_pc = err3d_dict

    # Scelta metodo migliore
    best_method = 'bilinear'
    if err3d_dict is not None:
        eb = err3d_dict['bilinear']; ec = err3d_dict['bicubic']
        eb_v = eb[~np.isnan(eb)] if eb.size else np.array([])
        ec_v = ec[~np.isnan(ec)] if ec.size else np.array([])
        m_bil = float(eb_v.mean()) if eb_v.size else float('inf')
        m_bic = float(ec_v.mean()) if ec_v.size else float('inf')
        best_method = 'bicubic' if m_bic < m_bil else 'bilinear'
        print(f"\n[Method] Best 3D lookup: {best_method.upper()}  "
              f"(bilinear={m_bil:.4f} mm, bicubic={m_bic:.4f} mm)")

    err3d_mm  = err3d_dict[best_method] if err3d_dict is not None else None
    coords_3d_h = (coords_3d_h_bic if best_method == 'bicubic'
                   else coords_3d_h_bil)
    if coords_3d_h is not None:
        print(f"\n  Suspicious points -> 3D (method={best_method}):")
        for i in range(len(coords_3d_h)):
            ph  = suspicious_pixels_hsi[i]
            xyz = coords_3d_h[i]
            print(f"  T{i}: HSI({ph[0]:.1f},{ph[1]:.1f}) -> "
                  f"X={xyz[0]:.3f} Y={xyz[1]:.3f} Z={xyz[2]:.3f} mm")

    # ── Step 7: Save Excel ────────────────────────────────────────────────────
    excel_path = os.path.join(output_dir, f'{sample_name}_registration_errors.xlsx')
    print(f"\n[Step 7] Saving Excel error report -> {excel_path}")
    save_excel_errors(
        err2d_px      = err2d_px,
        err3d_dict    = err3d_dict,
        data_hsi      = data_hsi,
        common_ids    = common_ids,
        marker_side_mm= marker_side_mm,
        output_path   = excel_path,
        err3d_dict_pc = err3d_dict_pc if dual else None,
        res_reg       = res_reg_actual,
        res_pc        = res_pc_actual  if dual else None,
    )

    # ── Step 8: Registration visualisation (render REGISTRAZIONE) ─────────────
    if save_images:
        print("\n[Step 8] Saving registration visualisation...")
        visualize_registration(
            render_rgb            = render_rgb_reg,
            hsi_2d                = hsi_2d,
            pts_render_h          = pts_render_h,
            data_hsi              = data_hsi,
            data_render           = data_render_reg,
            suspicious_pixels_hsi = suspicious_pixels_hsi,
            coords_3d_h           = coords_3d_h,
            output_path           = os.path.join(output_dir,
                                                 f'{sample_name}_registration.png'),
        )
    else:
        print("\n[Step 8] Skipped (save_images=False).")

    # ── Step 9+10: Spectral point cloud — streaming (build + export insieme) ────
    wavelengths = parse_wavelengths(meta)
    if wavelengths:
        print(f"  Wavelengths: {wavelengths[0]:.1f} ... {wavelengths[-1]:.1f} nm "
              f"({len(wavelengths)} bands)")

    if save_pointcloud:
        # La versione streaming non accumula mai l'intero array degli spettri in RAM.
        # Peak RAM ≈ pc_chunk_size × (3 + bands) × 4 byte  (es. 100K × 293 × 4 ≈ 117 MB).
        print("\n[Step 9] Building spectral point cloud...")
        base = os.path.join(output_dir, f'spectral_pointcloud_{sample_name}')
        valid_mask, n_valid = build_and_export_pointcloud_streaming(
            cube              = cube,
            xyz_map           = xyz_map_pc,
            H_inv             = H_pc_inv,
            output_ply_path   = (base + '.ply')  if export_ply_file else None,
            output_npz_prefix = base              if export_npy_file else None,
            wavelengths       = wavelengths,
            border_px         = border_px,
            reflectance_norm  = reflectance_norm,
            chunk_size        = pc_chunk_size,
        )
        # CSV opzionale: richiede un passaggio extra (campionamento casuale)
        if export_csv_file and n_valid > 0:
            print(f"\n[Step 10] Exporting CSV (max {csv_max_points:,} punti campionati)...")
            _export_csv_streaming(
                cube=cube, xyz_flat=xyz_map_pc.reshape(-1, 3),
                valid_idx=(np.where(valid_mask.ravel())[0]).astype(np.int32),
                H_inv32=H_pc_inv.astype(np.float32),
                W_r=xyz_map_pc.shape[1], cols_hsi=cube.shape[1], rows_hsi=cube.shape[0],
                wavelengths=wavelengths, output_path=base + '.csv',
                max_points=csv_max_points, reflectance_norm=reflectance_norm,
            )
    else:
        print("\n[Step 9+10] Point cloud skipped (save_pointcloud=False).")
        # Calcola solo valid_mask per la coverage map (se richiesta), senza scrivere file
        valid_mask = ~np.isnan(xyz_map_pc[:, :, 0])
        n_valid    = int(valid_mask.sum())

    # In modalità streaming i dati non vengono mantenuti in RAM
    points_xyz = None
    spectra    = None

    # ── Step 11: Coverage visualisation (render POINT CLOUD) ─────────────────
    if save_images:
        print("\n[Step 11] Saving coverage map...")
        visualize_coverage(
            render_rgb  = render_rgb_pc,
            valid_mask  = valid_mask,
            hsi_2d      = hsi_2d,
            H_inv       = H_pc_inv,
            output_path = os.path.join(output_dir, f'{sample_name}_coverage.png'),
        )
    else:
        print("\n[Step 11] Skipped (save_images=False).")

    # ── Summary ───────────────────────────────────────────────────────────────
    px_per_mm = _px_per_mm_from_data(data_hsi, marker_side_mm)
    print("\n" + "=" * 68)
    print("PIPELINE COMPLETED")
    print("=" * 68)
    print(f"  Res. registrazione : {res_reg_actual} mm/px")
    print(f"  Res. point cloud   : {res_pc_actual} mm/px")
    print(f"  2D error mean  : {err2d_px.mean():.4f} px  "
          f"= {err2d_px.mean()/px_per_mm:.4f} mm")
    print(f"  2D error max   : {err2d_px.max():.4f} px  "
          f"= {err2d_px.max()/px_per_mm:.4f} mm")
    if err3d_dict is not None:
        for tag, arr in (('bilinear', err3d_dict['bilinear']),
                         ('bicubic',  err3d_dict['bicubic'])):
            v = arr[~np.isnan(arr)]
            if v.size:
                best_lbl = ' <-- best' if tag == best_method else ''
                print(f"  3D REG {tag:<8s} mean: {v.mean():.4f} mm  "
                      f"median: {np.median(v):.4f} mm{best_lbl}")
    if dual and err3d_dict_pc is not None:
        for tag, arr in (('bilinear', err3d_dict_pc['bilinear']),
                         ('bicubic',  err3d_dict_pc['bicubic'])):
            v = arr[~np.isnan(arr)]
            if v.size:
                print(f"  3D PC  {tag:<8s} mean: {v.mean():.4f} mm  "
                      f"median: {np.median(v):.4f} mm")
    n_bands = len(wavelengths) if wavelengths else cube.shape[2]
    print(f"  Point cloud    : {n_valid:,} points  x {n_bands} bands")
    print(f"  Output dir     : {output_dir}")

    return {
        # Omografie
        'H': H, 'H_inv': H_inv,
        'H_pc': H_pc, 'H_pc_inv': H_pc_inv,
        # Registration
        'pts_hsi': pts_hsi, 'pts_render': pts_render_reg,
        'common_ids': common_ids,
        'data_hsi': data_hsi, 'data_render': data_render_reg,
        'err2d_px': err2d_px,
        'err3d_mm': err3d_mm,
        'err3d_dict': err3d_dict,
        'err3d_dict_pc': err3d_dict_pc if dual else None,
        'best_method': best_method,
        'pts_render_h': pts_render_h,
        'coords_3d_h': coords_3d_h,
        'coords_3d_h_bilinear': coords_3d_h_bil,
        'coords_3d_h_bicubic':  coords_3d_h_bic,
        # Render registrazione
        'render_rgb_reg': render_rgb_reg, 'depth_map_reg': depth_map_reg,
        'xyz_map_reg': xyz_map_reg, 'origin_xy_reg': origin_xy_reg,
        # Render point cloud  (alias 'render_rgb' per retrocompatibilità)
        'render_rgb': render_rgb_pc, 'depth_map': depth_map_pc,
        'xyz_map': xyz_map_pc, 'origin_xy': origin_xy_pc,
        # HSI
        'cube': cube, 'hsi_2d': hsi_2d, 'wavelengths': wavelengths,
        # Point cloud (streaming: xyz e spectra non sono in RAM, usa i file esportati)
        'xyz': points_xyz, 'spectra': spectra,
        'valid_mask': valid_mask, 'n_valid': n_valid,
        'bands': cube.shape[2],
    }


# =============================================================================
# USAGE EXAMPLE
# =============================================================================

if __name__ == "__main__":

    HSI_HDR    = r"C:\Users\leolu\Desktop\Polimi\Stage\SpectraBreast\Acquisizioni 23-04\SB019\SB019_001\SB019_raw.hdr"
    MESH_PATH  = r"C:\Users\leolu\Desktop\Polimi\Stage\SpectraBreast\SAMPLE1\surface_mesh.ply"
    ARUCO_JSON = r"C:\Users\leolu\Desktop\Polimi\Stage\SpectraBreast\SAMPLE1\aruco_markers_3d.json"
    OUTPUT_DIR = r"C:\Users\leolu\Desktop\Polimi\Stage\SpectraBreast\SAMPLE1\output_full"

    # ── Optional: suspicious pixels from HSI segmentation ─────────────────────
    hsi_rows, hsi_cols = 1024, 1280
    suspicious_mask = np.zeros((hsi_rows, hsi_cols), dtype=np.uint8)
    suspicious_mask[500:550, 500:550] = 1   # replace with real segmentation mask

    print('\n[Pre-step] Clustering suspicious pixels...')
    punti_sospetti, n_clusters = extract_suspicious_centroids(
        suspicious_mask, min_area=5
    )
    print(f'  {n_clusters} region(s) -> {n_clusters} centroid(s)\n')

    # ── Run ───────────────────────────────────────────────────────────────────
    result = run_full_pipeline(
        hsi_hdr_path             = HSI_HDR,
        mesh_path                = MESH_PATH,
        aruco_json_path          = ARUCO_JSON,
        output_dir               = OUTPUT_DIR,
        hsi_extraction_method    = 'mean',
        aruco_dict_type          = cv2.aruco.DICT_4X4_50,
        suspicious_pixels_hsi    = punti_sospetti,
        render_resolution_mm     = 0.3,   # fine — per il point cloud
        render_resolution_reg_mm = 0.5,   # grossolano — per l'omografia (12/12 inlier)
        render_margin_mm         = 10.0,
        chunk_size               = 500_000,
        marker_side_mm           = 6.8,
        use_subpix               = True,
        subpix_winsize           = 5,
        border_px                = 2,
        reflectance_norm         = True,
        export_ply_file          = True,
        export_npy_file          = True,
        export_csv_file          = False,
        sample_name              = 'SB019',
    )

    # ── Quick access ──────────────────────────────────────────────────────────
    print(f"\nPoint cloud : {result['xyz'].shape[0]:,} pts x "
          f"{result['spectra'].shape[1]} bands")
    if result['wavelengths']:
        print(f"Wavelengths : {result['wavelengths'][0]:.1f} ... "
              f"{result['wavelengths'][-1]:.1f} nm")
    if result['coords_3d_h'] is not None:
        print(f"\nSuspicious points -> 3D (best method = {result['best_method']}):")
        for i, xyz in enumerate(result['coords_3d_h']):
            print(f"  T{i}: X={xyz[0]:.3f}  Y={xyz[1]:.3f}  Z={xyz[2]:.3f} mm")
        # Confronto con l'altro metodo
        other = (result['coords_3d_h_bicubic']
                 if result['best_method'] == 'bilinear'
                 else result['coords_3d_h_bilinear'])
        if other is not None:
            other_name = ('bicubic' if result['best_method'] == 'bilinear' else 'bilinear')
            print(f"\nFor reference, same points via {other_name}:")
            for i, xyz in enumerate(other):
                print(f"  T{i}: X={xyz[0]:.3f}  Y={xyz[1]:.3f}  Z={xyz[2]:.3f} mm")